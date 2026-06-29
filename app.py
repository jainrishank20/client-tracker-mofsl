import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import json, os, io
from datetime import datetime, date
from pathlib import Path

# ── Google Sheets config ────────────────────────────────────────────────────
GSHEET_KEY  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gsheet_key.json")
GSHEET_ID   = "1RBaZYY8Eheet13UJy6eRMJIFUzU9Yii335l5x_H5KVo"

def sync_to_gsheet(trades: list):
    """Push all trade data to Google Sheets — 4 tabs with formatting, CMP, conditional colors."""
    import gspread
    from gspread_formatting import (
        CellFormat, Color, TextFormat, BooleanCondition, BooleanRule,
        ConditionalFormatRule, GridRange, format_cell_range, format_cell_ranges,
        set_frozen, get_conditional_format_rules, set_column_width
    )

    import time as _time
    from gspread.exceptions import APIError as _GAPIError

    # Rate limiter: max 45 API calls per minute (Google allows 60)
    # Rate limiter: enforce max 30 calls/min AND 1-second gap between calls.
    # Google allows 60 write req/min but also has a per-second burst limit.
    # Staying at 30/min with 1s spacing prevents both limits.
    _call_times = []
    _last_call  = [0.0]
    def _rate_limit():
        now = _time.time()
        gap = now - _last_call[0]
        if gap < 1.1:
            _time.sleep(1.1 - gap)
        _call_times[:] = [t for t in _call_times if now - t < 60]
        if len(_call_times) >= 30:
            wait = 62 - (now - _call_times[0])
            if wait > 0:
                _time.sleep(wait)
            _call_times.clear()
        _last_call[0] = _time.time()
        _call_times.append(_last_call[0])

    def _retry(fn, *args, **kwargs):
        _rate_limit()
        for attempt in range(6):
            try:
                return fn(*args, **kwargs)
            except _GAPIError as e:
                if e.response.status_code == 429:
                    _time.sleep(70)
                    _call_times.clear()
                else:
                    raise
        return fn(*args, **kwargs)

    # All formatting helpers go through _retry (which calls _rate_limit internally)
    _r_fcr  = format_cell_range
    _r_fcrs = format_cell_ranges
    _r_sf   = set_frozen
    _r_scw  = set_column_width
    def format_cell_range(*a, **kw):  return _retry(_r_fcr,  *a, **kw)
    def format_cell_ranges(*a, **kw): return _retry(_r_fcrs, *a, **kw)
    def set_frozen(*a, **kw):         return _retry(_r_sf,   *a, **kw)
    def set_column_width(*a, **kw):   return _retry(_r_scw,  *a, **kw)

    gc = gspread.service_account(filename=GSHEET_KEY)
    sh = gc.open_by_key(GSHEET_ID)

    # delete stale tabs from old versions
    for stale in ["⚡ Intraday", "📌 Script Summary"]:
        try:
            sh.del_worksheet(sh.worksheet(stale))
        except Exception:
            pass

    df_all = pd.DataFrame(trades)
    if df_all.empty:
        return "No trades to sync."

    df_all['entry_date'] = pd.to_datetime(df_all['entry_date'], errors='coerce')
    df_all['exit_date']  = pd.to_datetime(df_all['exit_date'],  errors='coerce')
    open_df   = df_all[df_all['exit_date'].isna()].copy()
    closed_df = df_all[df_all['exit_date'].notna()].copy()

    # Compute derived columns that build_dataframe() normally provides
    for _col in ["buy_price", "sell_price", "buy_qty", "sell_qty"]:
        closed_df[_col] = pd.to_numeric(closed_df[_col], errors="coerce").fillna(0)
        open_df[_col]   = pd.to_numeric(open_df[_col],   errors="coerce").fillna(0)
    closed_df["pnl"] = (closed_df["sell_price"] - closed_df["buy_price"]) * closed_df["buy_qty"]
    if "net_pnl" in closed_df.columns:
        closed_df["net_pnl"] = pd.to_numeric(closed_df["net_pnl"], errors="coerce").fillna(0)
    else:
        closed_df["net_pnl"] = closed_df["pnl"]
    open_df["invested"] = open_df["buy_price"] * open_df["buy_qty"]

    def total_charges(row, prefix):
        keys = [f'{prefix}_brokerage', f'{prefix}_stt', f'{prefix}_gst',
                f'{prefix}_stamp', f'{prefix}_txn_chrg']
        return sum(float(row.get(k, 0) or 0) for k in keys)

    # ── Formatting helpers ──────────────────────────────────────────────────────
    HEADER_BG   = Color(0.13, 0.13, 0.18)   # dark navy
    HEADER_FG   = Color(1, 1, 1)
    ALT_BG      = Color(0.95, 0.95, 0.97)
    GREEN_BG    = Color(0.84, 0.97, 0.89)
    RED_BG      = Color(0.99, 0.87, 0.87)
    GREEN_FG    = Color(0.08, 0.50, 0.25)
    RED_FG      = Color(0.65, 0.05, 0.05)

    header_fmt = CellFormat(
        backgroundColor=HEADER_BG,
        textFormat=TextFormat(bold=True, foregroundColor=HEADER_FG, fontSize=10),
        horizontalAlignment='CENTER'
    )
    num_fmt    = CellFormat(numberFormat={"type":"NUMBER","pattern":"#,##0.00"}, horizontalAlignment='RIGHT')
    center_fmt = CellFormat(horizontalAlignment='CENTER')

    def col_letter(n):
        """1-indexed column number to letter(s)."""
        s = ""
        while n:
            n, r = divmod(n-1, 26)
            s = chr(65+r) + s
        return s

    def upsert_ws(name, rows=200, cols=20):
        _time.sleep(1)
        try:
            ws = sh.worksheet(name)
            _retry(ws.clear)
            # Resize existing sheet if data exceeds current grid
            if ws.row_count < rows or ws.col_count < cols:
                ws.resize(rows=max(rows, ws.row_count), cols=max(cols, ws.col_count))
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title=name, rows=rows, cols=cols)
        # clear all cell formatting so stale styles (e.g. old TOTAL row) don't persist
        try:
            sh.batch_update({'requests': [{'updateCells': {
                'range': {'sheetId': ws.id},
                'fields': 'userEnteredFormat'
            }}]})
        except Exception:
            pass
        return ws

    def write_df(ws, df):
        if df.empty:
            ws.update([["No data"]], value_input_option='RAW')
            return
        df2 = df.copy()
        for c in df2.select_dtypes(include=['datetime64[ns]']).columns:
            df2[c] = df2[c].dt.strftime('%Y-%m-%d').fillna('')
        df2 = df2.fillna('').astype(str)
        # convert numeric-looking strings back to numbers
        def try_num(v):
            try: return float(v) if '.' in v else int(v)
            except: return v
        data = [df2.columns.tolist()] + [[try_num(v) for v in r] for r in df2.values.tolist()]
        _retry(ws.update, data, value_input_option='USER_ENTERED')
        _time.sleep(0.5)

    def apply_header_fmt(ws, ncols):
        fmt_range = f"A1:{col_letter(ncols)}1"
        format_cell_range(ws, fmt_range, header_fmt)
        set_frozen(ws, rows=1)

    _cleared_ws_rules = set()
    _pending_rules = {}  # ws.id -> rules object, flushed by flush_pnl_rules()

    def apply_pnl_conditional(ws, pnl_col_idx, nrows):
        """Queue Green/Red conditional rules — call flush_pnl_rules(ws) to save in one API call."""
        col = col_letter(pnl_col_idx)
        rng = GridRange.from_a1_range(f"{col}2:{col}{nrows+1}", ws)
        if ws.id not in _pending_rules:
            rules = get_conditional_format_rules(ws)
            rules.clear()
            _pending_rules[ws.id] = rules
        rules = _pending_rules[ws.id]
        rules.append(ConditionalFormatRule(
            ranges=[rng],
            booleanRule=BooleanRule(
                condition=BooleanCondition('NUMBER_GREATER', ['0']),
                format=CellFormat(backgroundColor=GREEN_BG,
                                  textFormat=TextFormat(bold=True, foregroundColor=GREEN_FG))
            )
        ))
        rules.append(ConditionalFormatRule(
            ranges=[rng],
            booleanRule=BooleanRule(
                condition=BooleanCondition('NUMBER_LESS', ['0']),
                format=CellFormat(backgroundColor=RED_BG,
                                  textFormat=TextFormat(bold=True, foregroundColor=RED_FG))
            )
        ))

    def flush_pnl_rules(ws):
        """Save all queued conditional rules for a worksheet in a single API call."""
        if ws.id in _pending_rules:
            _pending_rules[ws.id].save()
            del _pending_rules[ws.id]

    def apply_num_cols(ws, col_indices, nrows):
        # batch all column formats into a single API call
        ranges = [(f"{col_letter(i)}2:{col_letter(i)}{nrows+1}", num_fmt) for i in col_indices]
        if ranges:
            format_cell_ranges(ws, ranges)

    # ── Build Google Finance formula map ────────────────────────────────────────
    def gfinance_formula(script):
        """Returns =GOOGLEFINANCE("NSE:TICKER","price") formula for a script."""
        overrides = load_ticker_overrides()
        if script in overrides and overrides[script]:
            sym = overrides[script].strip().upper().replace(".NS","")
        else:
            sym = SYMBOL_MAP.get(script, script.replace(" ","").replace("&","").replace(".",""))
        return f'=GOOGLEFINANCE("NSE:{sym}","price")'

    # also keep python CMP for overview unrealized calc
    cmp_map = {}
    if not open_df.empty:
        scripts = tuple(open_df['script'].unique())
        try:
            cmp_map = fetch_cmp(scripts)
        except Exception:
            cmp_map = {}

    pct_fmt  = CellFormat(numberFormat={"type":"NUMBER","pattern":'0.00"%"'}, horizontalAlignment='RIGHT')
    tot_fmt  = CellFormat(backgroundColor=Color(0.13,0.13,0.18),
                          textFormat=TextFormat(bold=True, foregroundColor=Color(1,1,1)),
                          horizontalAlignment='RIGHT')
    synced_at = datetime.now(tz=__import__('zoneinfo').ZoneInfo('Asia/Kolkata')).strftime('%d %b %Y %I:%M %p IST')

    def add_totals_row(ws, df, num_cols, pct_cols, last_row):
        """Append a TOTALS row summing numeric columns."""
        ncols = len(df.columns)
        tot_row = [''] * ncols
        tot_row[0] = 'TOTAL'
        for ci in num_cols:
            col = col_letter(ci)
            tot_row[ci-1] = f'=SUM({col}2:{col}{last_row})'
        ws.update([tot_row], range_name=f'A{last_row+1}:{col_letter(ncols)}{last_row+1}',
                  value_input_option='USER_ENTERED')
        format_cell_range(ws, f'A{last_row+1}:{col_letter(ncols)}{last_row+1}', tot_fmt)
        for ci in pct_cols:
            col = col_letter(ci)
            format_cell_range(ws, f'{col}{last_row+1}', pct_fmt)

    # ═══════════════════════════════════════════════════════════════════════════
    # TAB 1 — Overview  (+ last synced timestamp)
    # ═══════════════════════════════════════════════════════════════════════════
    rows_ov = []
    ov_errors = []
    for c in CLIENTS:
        try:
            cd = closed_df[closed_df['client']==c]
            od = open_df[open_df['client']==c]
            gross_c = ((cd['sell_price'] - cd['buy_price']) * cd['sell_qty']).sum() if not cd.empty else 0
            buy_chg = cd.apply(lambda r: total_charges(r,'buy'), axis=1).sum()  if not cd.empty else 0
            sel_chg = cd.apply(lambda r: total_charges(r,'sell'),axis=1).sum() if not cd.empty else 0
            net_c   = gross_c - buy_chg - sel_chg
            unreal  = 0
            if not od.empty:
                for _, r in od.iterrows():
                    cmp = cmp_map.get(r['script'])
                    if cmp and cmp not in ('', None):
                        unreal += (float(cmp) - float(r['buy_price'])) * float(r['buy_qty'])
            rows_ov.append({
                'Client Code':        c,
                'Client Name':        CLIENT_NAMES.get(c, c),
                'Open Positions':     int(open_df[open_df['client']==c]['script'].nunique()) if not od.empty else 0,
                'Closed Trades':      len(cd),
                'Realized Gross P&L': round(gross_c, 2),
                'Total Charges':      round(buy_chg + sel_chg, 2),
                'Realized Net P&L':   round(net_c, 2),
                'Unrealized P&L':     round(unreal, 2),
                'Total P&L':          round(net_c + unreal, 2),
            })
        except Exception as _ov_err:
            ov_errors.append(f"{c}: {_ov_err}")
    df_ov = pd.DataFrame(rows_ov)
    ws_ov = upsert_ws("📊 Overview", rows=25, cols=12)
    write_df(ws_ov, df_ov)
    apply_header_fmt(ws_ov, len(df_ov.columns))
    apply_num_cols(ws_ov, [5,6,7,8,9], len(df_ov))
    for ci in [7,8,9]: apply_pnl_conditional(ws_ov, ci, len(df_ov))
    flush_pnl_rules(ws_ov)
    add_totals_row(ws_ov, df_ov, [5,6,7,8,9], [], len(df_ov)+1)
    # last synced timestamp
    ws_ov.update([[f'Last synced: {synced_at}']],
                 range_name=f'A{len(df_ov)+3}', value_input_option='USER_ENTERED')
    format_cell_range(ws_ov, f'A{len(df_ov)+3}',
                      CellFormat(textFormat=TextFormat(italic=True, foregroundColor=Color(0.5,0.5,0.5))))
    set_column_width(ws_ov, 'A', 110); set_column_width(ws_ov, 'B', 140)

    # ═══════════════════════════════════════════════════════════════════════════
    # TAB 2 — Open Positions  (consolidated per script+client, sorted by % Return)
    # ═══════════════════════════════════════════════════════════════════════════
    _time.sleep(2)
    if not open_df.empty:
        grp = open_df.groupby(['client','script']).apply(lambda g: pd.Series({
            'Qty':           g['buy_qty'].sum(),
            'Avg Buy Price': round((g['buy_price']*g['buy_qty']).sum() / g['buy_qty'].sum(), 2),
            'Invested (₹)': round((g['buy_price']*g['buy_qty']).sum(), 2),
            'First Entry':   g['entry_date'].min().strftime('%Y-%m-%d'),
            'Last Entry':    g['entry_date'].max().strftime('%Y-%m-%d'),
        })).reset_index()
        grp['Client Name'] = grp['client'].map(CLIENT_NAMES)
        grp['Days Held']   = (pd.Timestamp.today() - pd.to_datetime(grp['First Entry'])).dt.days.astype(int)
        grp['CMP']         = grp['script'].map(gfinance_formula)
        show_op = grp[['client','Client Name','script','Qty','Avg Buy Price',
                        'Invested (₹)','CMP','First Entry','Last Entry','Days Held']].copy()
        show_op.columns = ['Client','Client Name','Script','Qty','Avg Buy Price',
                           'Invested (₹)','CMP','First Entry','Last Entry','Days Held']
        # compute % return for sorting using Python-fetched CMP (cmp_map has real numbers; CMP col has formulas)
        def _pct(row):
            cmp = cmp_map.get(row['Script'])
            try:
                return (float(cmp) - float(row['Avg Buy Price'])) / float(row['Avg Buy Price']) * 100
            except Exception:
                return float('-inf')
        show_op['_sort_pct'] = show_op.apply(_pct, axis=1)
        show_op = show_op.sort_values('_sort_pct', ascending=False).drop(columns='_sort_pct').reset_index(drop=True)
    else:
        show_op = pd.DataFrame(columns=['Client','Client Name','Script','Qty','Avg Buy Price',
                                        'Invested (₹)','CMP','First Entry','Last Entry','Days Held'])
    ws_op = upsert_ws("📋 Open Positions", rows=max(len(show_op)+15,100), cols=15)
    write_df(ws_op, show_op)
    nrows_op = len(show_op)
    ncols_op = len(show_op.columns)
    apply_header_fmt(ws_op, ncols_op)
    # formula columns: K=Unrealized P&L  L=% Return  (CMP=G col7, Qty=D col4, AvgBuy=E col5)
    if nrows_op > 0:
        ws_op.update([['Unrealized P&L','% Return']],
                     range_name='K1:L1', value_input_option='USER_ENTERED')
        format_cell_range(ws_op, 'K1:L1', header_fmt)
        fml = []
        for ri in range(2, nrows_op+2):
            g,d,e = f'G{ri}',f'D{ri}',f'E{ri}'
            fml.append([f'=IF(ISNUMBER({g}),({g}-{e})*{d},"")',
                        f'=IF(ISNUMBER({g}),({g}-{e})/{e}*100,"")'])
        ws_op.update(fml, range_name=f'K2:L{nrows_op+1}', value_input_option='USER_ENTERED')
        apply_num_cols(ws_op, [5,6,11], nrows_op)
        format_cell_range(ws_op, f'L2:L{nrows_op+1}', pct_fmt)
        apply_pnl_conditional(ws_op, 11, nrows_op)
        apply_pnl_conditional(ws_op, 12, nrows_op)
        flush_pnl_rules(ws_op)
        tr = nrows_op + 2  # totals row number
        tot = ['TOTAL','','','','','','','','','',
               f'=SUM(K2:K{nrows_op+1})',
               f'=IF(F{tr}<>0,K{tr}/F{tr}*100,"")']  # overall % = unrealized / invested
        tot[3] = f'=SUM(D2:D{nrows_op+1})'   # Qty
        tot[5] = f'=SUM(F2:F{nrows_op+1})'   # Invested
        ws_op.update([tot], range_name=f'A{tr}:L{tr}', value_input_option='USER_ENTERED')
        format_cell_range(ws_op, f'A{tr}:L{tr}', tot_fmt)
        format_cell_range(ws_op, f'L{tr}', pct_fmt)
    set_frozen(ws_op, rows=1, cols=3)

    # ═══════════════════════════════════════════════════════════════════════════
    # TAB 3 — Script Summary  (cross-client exposure per script)
    # ═══════════════════════════════════════════════════════════════════════════
    _time.sleep(2)
    # build master ledger — all trades open + closed, sortable/filterable in Sheets
    ledger_rows = []
    for _, r in df_all.iterrows():
        is_open = pd.isna(r['exit_date'])
        gross   = round((float(r['sell_price']) - float(r['buy_price'])) * float(r['sell_qty']), 2) if not is_open else ''
        charges = round(total_charges(r,'buy') + total_charges(r,'sell'), 2)
        net     = round(gross - charges, 2) if gross != '' else ''
        pct     = round(gross / (float(r['buy_price']) * float(r['sell_qty'])) * 100, 2) if gross != '' and float(r['buy_price']) > 0 else ''
        hold    = (r['exit_date'] - r['entry_date']).days if not is_open else ''
        ledger_rows.append({
            'Client':       r['client'],
            'Client Name':  CLIENT_NAMES.get(r['client'], r['client']),
            'Script':       r['script'],
            'Status':       'Open' if is_open else 'Closed',
            'Entry Date':   r['entry_date'].strftime('%Y-%m-%d'),
            'Buy Qty':      float(r['buy_qty']),
            'Buy Price':    float(r['buy_price']),
            'Invested (₹)': round(float(r['buy_price']) * float(r['buy_qty']), 2),
            'Exit Date':    '' if is_open else r['exit_date'].strftime('%Y-%m-%d'),
            'Sell Qty':     float(r['sell_qty']) if not is_open else '',
            'Sell Price':   float(r['sell_price']) if not is_open else '',
            'Gross P&L':    gross,
            'Charges':      charges,
            'Net P&L':      net,
            '% Return':     pct,
            'Hold Days':    hold,
        })
    df_ledger = pd.DataFrame(ledger_rows).sort_values(['Client','Script','Entry Date']).reset_index(drop=True)
    ws_sc = upsert_ws("📌 Trade Ledger", rows=max(len(df_ledger)+10, 200), cols=17)
    write_df(ws_sc, df_ledger)
    nrows_sc = len(df_ledger)
    apply_header_fmt(ws_sc, len(df_ledger.columns))
    apply_num_cols(ws_sc, [7,8,12,13,14], nrows_sc)
    format_cell_range(ws_sc, f'O2:O{nrows_sc+1}', pct_fmt)
    apply_pnl_conditional(ws_sc, 12, nrows_sc)
    apply_pnl_conditional(ws_sc, 14, nrows_sc)
    flush_pnl_rules(ws_sc)
    set_frozen(ws_sc, rows=1, cols=3)

    # ═══════════════════════════════════════════════════════════════════════════
    # TAB 4 — P&L Summary  (per client: booked, unrealized, total)
    # ═══════════════════════════════════════════════════════════════════════════
    _time.sleep(2)
    pnl_rows = []
    pnl_errors = []
    for c in CLIENTS:
        try:
            cd = closed_df[closed_df['client']==c]
            od = open_df[open_df['client']==c]
            # booked
            gross_c   = ((cd['sell_price']-cd['buy_price'])*cd['sell_qty']).sum() if not cd.empty else 0
            buy_chg   = cd.apply(lambda r: total_charges(r,'buy'), axis=1).sum()  if not cd.empty else 0
            sel_chg   = cd.apply(lambda r: total_charges(r,'sell'),axis=1).sum()  if not cd.empty else 0
            net_booked= gross_c - buy_chg - sel_chg
            total_chg = buy_chg + sel_chg
            # unrealized (python CMP for reference; sheet has GOOGLEFINANCE live)
            unreal_py = 0
            if not od.empty:
                for _, r in od.iterrows():
                    cmp = cmp_map.get(r['script'])
                    if cmp:
                        unreal_py += (float(cmp)-float(r['buy_price']))*float(r['buy_qty'])
            capital   = round((od['buy_price']*od['buy_qty']).sum(), 2) if not od.empty else 0
            # win rate
            if not cd.empty:
                wins     = ((cd['sell_price']-cd['buy_price'])>0).sum()
                win_rate = round(wins/len(cd)*100, 1)
                best     = round(((cd['sell_price']-cd['buy_price'])*cd['sell_qty']).max(), 2)
                worst    = round(((cd['sell_price']-cd['buy_price'])*cd['sell_qty']).min(), 2)
                trades_count = len(cd)
            else:
                win_rate = trades_count = best = worst = 0
            pnl_rows.append({
                'Client':            c,
                'Name':              CLIENT_NAMES.get(c, c),
                'Open Positions':    len(od['script'].unique()) if not od.empty else 0,
                'Capital Deployed':  capital,
                'Closed Trades':     trades_count,
                'Booked Gross P&L':  round(gross_c, 2),
                'Total Charges':     round(total_chg, 2),
                'Booked Net P&L':    round(net_booked, 2),
                'Unrealized P&L*':   round(unreal_py, 2),
                'Total P&L*':        round(net_booked + unreal_py, 2),
                'Win Rate %':        win_rate,
                'Best Trade':        best,
                'Worst Trade':       worst,
            })
        except Exception as _pnl_err:
            pnl_errors.append(f"{c}: {_pnl_err}")
    df_pnl = pd.DataFrame(pnl_rows)
    ws_pnl = upsert_ws("💰 P&L Summary", rows=25, cols=14)
    write_df(ws_pnl, df_pnl)
    apply_header_fmt(ws_pnl, len(df_pnl.columns))
    nrows_pnl = len(df_pnl)
    apply_num_cols(ws_pnl, [4,6,7,8,9,10,12,13], nrows_pnl)
    apply_pnl_conditional(ws_pnl, 8,  nrows_pnl)
    apply_pnl_conditional(ws_pnl, 9,  nrows_pnl)
    apply_pnl_conditional(ws_pnl, 10, nrows_pnl)
    apply_pnl_conditional(ws_pnl, 12, nrows_pnl)
    apply_pnl_conditional(ws_pnl, 13, nrows_pnl)
    flush_pnl_rules(ws_pnl)
    add_totals_row(ws_pnl, df_pnl, [4,6,7,8,9,10,12,13], [11], nrows_pnl+1)
    # footnote: * = based on last fetched CMP, not live in this tab
    ws_pnl.update([['* Unrealized & Total P&L based on last CMP fetch. See Open Positions tab for live values.']],
                  range_name=f'A{nrows_pnl+3}', value_input_option='USER_ENTERED')
    format_cell_range(ws_pnl, f'A{nrows_pnl+3}',
                      CellFormat(textFormat=TextFormat(italic=True, foregroundColor=Color(0.5,0.5,0.5))))
    set_frozen(ws_pnl, rows=1, cols=2)

    # ═══════════════════════════════════════════════════════════════════════════
    # TAB 5 — Closed Trades
    # ═══════════════════════════════════════════════════════════════════════════
    _time.sleep(2)
    if not closed_df.empty:
        cl = closed_df.copy()
        cl['Client Name']  = cl['client'].map(CLIENT_NAMES)
        cl['Gross P&L']    = ((cl['sell_price'] - cl['buy_price']) * cl['sell_qty']).round(2)
        cl['Buy Charges']  = cl.apply(lambda r: round(total_charges(r,'buy'), 2), axis=1)
        cl['Sell Charges'] = cl.apply(lambda r: round(total_charges(r,'sell'),2), axis=1)
        cl['Net P&L']      = (cl['Gross P&L'] - cl['Buy Charges'] - cl['Sell Charges']).round(2)
        cl['% Return']     = (cl['Gross P&L'] / (cl['buy_price'] * cl['sell_qty']) * 100).round(2)
        cl['Type']         = cl.apply(lambda r: 'Intraday'
                              if r['entry_date'].date()==r['exit_date'].date() else 'Delivery', axis=1)
        cl['Hold Days']    = (cl['exit_date'] - cl['entry_date']).dt.days
        cl['Entry Date']   = cl['entry_date'].dt.strftime('%Y-%m-%d')
        cl['Exit Date']    = cl['exit_date'].dt.strftime('%Y-%m-%d')
        show_cl = cl[['client','Client Name','script','sell_qty','buy_price','sell_price',
                       'Entry Date','Exit Date','Hold Days','Gross P&L','Buy Charges',
                       'Sell Charges','Net P&L','% Return','Type']].copy()
        show_cl.columns = ['Client','Client Name','Script','Qty','Buy Price','Sell Price',
                           'Entry Date','Exit Date','Hold Days','Gross P&L','Buy Charges',
                           'Sell Charges','Net P&L','% Return','Type']
        show_cl = show_cl.sort_values(['Client','Exit Date'], ascending=[True,False]).reset_index(drop=True)
    else:
        show_cl = pd.DataFrame()
    ws_cl = upsert_ws("📈 Closed Trades", rows=max(len(show_cl)+10,100), cols=16)
    write_df(ws_cl, show_cl)
    apply_header_fmt(ws_cl, len(show_cl.columns) if not show_cl.empty else 15)
    if not show_cl.empty:
        nrows_cl = len(show_cl)
        apply_num_cols(ws_cl, [5,6,10,11,12,13], nrows_cl)
        format_cell_range(ws_cl, f'N2:N{nrows_cl+1}', pct_fmt)
        apply_pnl_conditional(ws_cl, 10, nrows_cl)
        apply_pnl_conditional(ws_cl, 13, nrows_cl)
        apply_pnl_conditional(ws_cl, 14, nrows_cl)
        flush_pnl_rules(ws_cl)
        add_totals_row(ws_cl, show_cl, [10,11,12,13], [14], nrows_cl+1)
    set_frozen(ws_cl, rows=1, cols=3)

    # ═══════════════════════════════════════════════════════════════════════════
    # TAB 6 — Monthly P&L
    # ═══════════════════════════════════════════════════════════════════════════
    _time.sleep(1)
    if not closed_df.empty:
        _mn_df = closed_df.copy()
        _mn_df["Month"] = _mn_df["exit_date"].dt.to_period("M").astype(str)
        _mn_rows = []
        for _month, _mg in _mn_df.groupby("Month"):
            _row = {"Month": _month}
            _row["All Clients Gross"] = round(_mg["pnl"].sum(), 2)
            _row["All Clients Net"]   = round(_mg["net_pnl"].sum(), 2)
            _row["Trades"]            = len(_mg)
            for _c in CLIENTS:
                _cg = _mg[_mg["client"] == _c]
                _row[f"{CLIENT_NAMES[_c]} Net"] = round(_cg["net_pnl"].sum(), 2) if not _cg.empty else 0
            _mn_rows.append(_row)
        df_monthly = pd.DataFrame(_mn_rows).sort_values("Month", ascending=False)
        ws_mn = upsert_ws("📆 Monthly P&L", rows=max(len(df_monthly)+20, 50),
                          cols=len(df_monthly.columns)+2)
        write_df(ws_mn, df_monthly)
        apply_header_fmt(ws_mn, len(df_monthly.columns))
        _mn_num_cols = list(range(2, 4 + len(CLIENTS) + 1))
        apply_num_cols(ws_mn, _mn_num_cols, len(df_monthly))
        for _cc in range(2, 4 + len(CLIENTS) + 1):
            apply_pnl_conditional(ws_mn, _cc, len(df_monthly))
        flush_pnl_rules(ws_mn)
        add_totals_row(ws_mn, df_monthly, _mn_num_cols, [], len(df_monthly)+1)
        set_frozen(ws_mn, rows=1, cols=1)

    # ═══════════════════════════════════════════════════════════════════════════
    # TAB 7 — Capital (Payin / Payout)
    # ═══════════════════════════════════════════════════════════════════════════
    _time.sleep(1)
    try:
        import os as _os
        _ledger_file = _os.path.join(_os.path.dirname(__file__), "ledger.json")
        _ledger_data = json.loads(open(_ledger_file).read()) if _os.path.exists(_ledger_file) else {}
    except Exception:
        _ledger_data = {}

    cap_rows_gs = []
    for c in CLIENTS:
        entries = _ledger_data.get(c, [])
        payin  = sum(e["amount"] for e in entries if e["amount"] > 0)
        payout = sum(-e["amount"] for e in entries if e["amount"] < 0)
        od_c   = open_df[open_df["client"] == c] if not open_df.empty else pd.DataFrame()
        deployed = (od_c["buy_price"] * od_c["buy_qty"]).sum() if not od_c.empty else 0
        net_closed_pnl = closed_df[closed_df["client"]==c]["net_pnl"].sum() if not closed_df.empty else 0
        cap_rows_gs.append({
            "Client Code":    c,
            "Client Name":    CLIENT_NAMES.get(c, c),
            "Total Payin":    round(payin, 2),
            "Total Payout":   round(payout, 2),
            "Net Capital":    round(payin - payout, 2),
            "Deployed (Open)":round(deployed, 2),
            "Booked Net P&L": round(net_closed_pnl, 2),
            "Entries":        len(entries),
        })
    df_cap_gs = pd.DataFrame(cap_rows_gs)
    ws_cap = upsert_ws("💰 Capital", rows=max(len(df_cap_gs)+20, 50), cols=10)
    write_df(ws_cap, df_cap_gs)
    apply_header_fmt(ws_cap, len(df_cap_gs.columns))
    if not df_cap_gs.empty:
        apply_num_cols(ws_cap, [3,4,5,6,7], len(df_cap_gs))
        apply_pnl_conditional(ws_cap, 7, len(df_cap_gs))
        flush_pnl_rules(ws_cap)
        add_totals_row(ws_cap, df_cap_gs, [3,4,5,6,7], [], len(df_cap_gs)+1)

    # Individual ledger entries tab
    all_ledger_rows = []
    for c in CLIENTS:
        for e in _ledger_data.get(c, []):
            all_ledger_rows.append({
                "Client":    c,
                "Name":      CLIENT_NAMES.get(c, c),
                "Date":      e["date"],
                "Type":      e["type"],
                "Ledger":    e.get("ledger", ""),
                "Amount":    e["amount"],
                "Narration": e.get("narration", ""),
            })
    if all_ledger_rows:
        df_led_gs = pd.DataFrame(all_ledger_rows).sort_values(["Client","Date"], ascending=[True,False])
        ws_led = upsert_ws("📒 Ledger Entries", rows=max(len(df_led_gs)+10, 100), cols=8)
        write_df(ws_led, df_led_gs)
        apply_header_fmt(ws_led, len(df_led_gs.columns))
        apply_num_cols(ws_led, [6], len(df_led_gs))
        apply_pnl_conditional(ws_led, 6, len(df_led_gs))
        flush_pnl_rules(ws_led)

    total = len(open_df) + len(closed_df)
    all_errors = (ov_errors or []) + (pnl_errors or [])
    err_str = f" ⚠️ Errors: {'; '.join(all_errors)}" if all_errors else ""
    return f"✅ Synced {total} trades ({len(open_df)} open, {len(closed_df)} closed) · {len(rows_ov)} clients · Last synced: {synced_at}{err_str}"

# ── CONFIG ────────────────────────────────────────────────────────────────────
DATA_FILE    = os.path.join(os.path.dirname(__file__), "trades.json")
LEDGER_FILE  = os.path.join(os.path.dirname(__file__), "ledger.json")
CLIENTS      = ["RIMK1209","RIMK1220","RIMK1238","RIMK1248","RIMK1249","RIMK1252"]
CLIENT_NAMES = {"RIMK1209":"Sathyavrath","RIMK1220":"Kalpana",
                "RIMK1238":"Iranna","RIMK1248":"Udayakumar",
                "RIMK1249":"Sundareshwari","RIMK1252":"Savitha"}
TELEGRAM_TOKEN = ""
TELEGRAM_CHAT  = ""

st.set_page_config(page_title="Client Tracker MOFSL", page_icon="📈",
                   layout="wide", initial_sidebar_state="collapsed")

# ── AUTH ──────────────────────────────────────────────────────────────────────
import bcrypt as _bcrypt

_USERS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "users.json")

def _load_users():
    if not os.path.exists(_USERS_FILE): return {}
    with open(_USERS_FILE) as f: return json.load(f)

def _check_login(username, password):
    users = _load_users()
    u = users.get(username)
    if not u: return False
    return _bcrypt.checkpw(password.encode(), u["password_hash"].encode())

def _get_role(username):
    return _load_users().get(username, {}).get("role", "viewer")

if "auth_ok" not in st.session_state:
    st.session_state.auth_ok = False
    st.session_state.auth_user = ""
    st.session_state.auth_role = ""

if not st.session_state.auth_ok:
    st.markdown("""
    <style>
    #MainMenu {visibility:hidden;} footer {visibility:hidden;}
    .block-container {max-width:420px; padding-top:80px;}
    </style>""", unsafe_allow_html=True)
    st.markdown("## 📈 Client Tracker MOFSL")
    st.markdown("Please log in to continue.")
    with st.form("login_form"):
        _u = st.text_input("Username")
        _p = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login", use_container_width=True, type="primary")
    if submitted:
        if _check_login(_u, _p):
            st.session_state.auth_ok   = True
            st.session_state.auth_user = _u
            st.session_state.auth_role = _get_role(_u)
            st.rerun()
        else:
            st.error("Invalid username or password.")
    st.stop()

_IS_ADMIN = (st.session_state.auth_role == "admin")

st.markdown("""
<style>
/* ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   DESIGN TOKENS
   bg0  = page background
   bg1  = card/surface
   bg2  = elevated card
   bdr  = border colour
   acc  = accent (indigo)
   txt  = primary text
   muted= secondary text
   up   = profit green
   dn   = loss red
   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ */

/* ── reset & base ── */
html, body,
[data-testid="stAppViewContainer"],
[data-testid="stMain"] {
    background: #0f1117 !important;
    color: #e8eaf0 !important;
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
}
[data-testid="stSidebar"],
[data-testid="collapsedControl"],
[data-testid="stDecoration"]  { display:none !important; }
/* keep header visible for theme toggle, just reduce its height */
[data-testid="stHeader"] { background:transparent !important; }
.block-container {
    padding-top: 0 !important;
    padding-left: 2.5rem !important;
    padding-right: 2.5rem !important;
    max-width: 100% !important;
}
* { box-sizing: border-box; }

/* ── navbar ── */
.brand-bar {
    background: #080a10;
    border-bottom: 1px solid #1f2232;
    padding: 0 28px;
    display: flex;
    align-items: center;
    height: 52px;
    margin: 0 -2.5rem;
    position: sticky; top:0; z-index:100;
}
.brand-bar-title { font-size:1rem; font-weight:800; color:#fff; letter-spacing:-.01em; }
.brand-bar-title em { color:#818cf8; font-style:normal; }


/* ── client pills ── */
div[data-testid="stPills"] { margin: 10px 0 4px 0 !important; }
[data-testid="stPillsItem"] {
    border-radius: 6px !important;
    font-size: .78rem !important; font-weight: 600 !important;
    padding: 4px 14px !important;
    border: 1px solid #1f2232 !important;
    color: #9ca3af !important;
    background: #131520 !important;
    transition: all .15s !important;
}
[data-testid="stPillsItem"]:hover { border-color:#818cf8 !important; color:#fff !important; }
[data-testid="stPillsItem"][aria-selected="true"] {
    background: #312e81 !important;
    border-color: #818cf8 !important;
    color: #c7d2fe !important;
}

/* ── metric cards ── */
[data-testid="stMetric"] {
    background: #131520;
    border: 1px solid #1f2232;
    border-radius: 10px;
    padding: 18px 20px !important;
}
[data-testid="stMetricValue"] {
    font-size: 1.55rem !important; font-weight: 800 !important;
    color: #f1f5f9 !important; letter-spacing: -.03em;
}
[data-testid="stMetricLabel"] {
    font-size: .68rem !important; font-weight: 700 !important;
    text-transform: uppercase; letter-spacing: .09em;
    color: #6b7280 !important;
}
[data-testid="stMetricDelta"] svg { display:none; }
[data-testid="stMetricDelta"] > div {
    font-size: .75rem !important; font-weight: 500 !important;
    color: #9ca3af !important;
}

/* ── section labels ── */
.kpi-head {
    font-size: .65rem; font-weight: 700; letter-spacing: .12em;
    text-transform: uppercase; color: #818cf8;
    margin: 24px 0 10px 2px;
}
.section-head {
    font-size: .68rem; font-weight: 700; letter-spacing: .1em;
    text-transform: uppercase; color: #4b5563;
    margin: 16px 0 8px 0;
}

/* ── content tabs ── */
[data-testid="stTabs"] [role="tablist"] {
    border-bottom: 1px solid #1f2232 !important;
    gap: 0 !important;
}
[data-testid="stTabs"] button {
    font-size: .82rem !important; font-weight: 600 !important;
    color: #6b7280 !important;
    border: none !important; border-radius: 0 !important;
    padding: 8px 20px !important;
    border-bottom: 2px solid transparent !important;
}
[data-testid="stTabs"] button[aria-selected="true"] {
    color: #fff !important;
    border-bottom: 2px solid #818cf8 !important;
    background: transparent !important;
}
[data-testid="stTabs"] button:hover { color: #e8eaf0 !important; }

/* ── dataframes ── */
div[data-testid="stDataFrameContainer"] {
    border: 1px solid #1f2232 !important;
    border-radius: 8px !important;
    overflow: hidden;
}
div[data-testid="stDataFrameContainer"] table td {
    font-size: .8rem !important; padding: 7px 12px !important;
    color: #d1d5db !important;
    border-bottom: 1px solid #1a1d2e !important;
}
div[data-testid="stDataFrameContainer"] table th {
    font-size: .68rem !important; text-transform: uppercase;
    letter-spacing: .07em; color: #6b7280 !important;
    background: #0d0f1a !important;
    border-bottom: 1px solid #1f2232 !important;
    padding: 9px 12px !important;
}

/* ── buttons ── */
[data-testid="stButton"] > button {
    border-radius: 7px !important;
    font-weight: 600 !important; font-size: .82rem !important;
    border: 1px solid #1f2232 !important;
    background: #131520 !important; color: #d1d5db !important;
    padding: 6px 16px !important;
    transition: all .15s !important;
}
[data-testid="stButton"] > button:hover {
    border-color: #818cf8 !important; color: #fff !important;
    background: #1e1b4b !important;
}
[data-testid="stButton"] > button[kind="primary"] {
    background: #312e81 !important; border-color: #818cf8 !important;
    color: #c7d2fe !important;
}

/* ── inputs ── */
input, textarea,
[data-baseweb="select"] > div,
[data-baseweb="input"]  > div  {
    background: #0d0f1a !important;
    border: 1px solid #1f2232 !important;
    border-radius: 7px !important;
    color: #e8eaf0 !important;
    font-size: .83rem !important;
}

/* ── dividers ── */
hr { border-color: #1f2232 !important; margin: 14px 0 !important; }

/* ── P&L HTML colours ── */
.up   { color: #34d399 !important; font-weight: 700; }
.down { color: #f87171 !important; font-weight: 700; }
.flat { color: #6b7280; }

/* ── client cards ── */
.client-card {
    background: #131520;
    border: 1px solid #1f2232;
    border-radius: 10px;
    padding: 16px 18px;
    text-align: center;
    height: 100%;
}
</style>
""", unsafe_allow_html=True)

# ── DATA ─────────────────────────────────────────────────────────────────────
def load():
    if not os.path.exists(DATA_FILE): return []
    with open(DATA_FILE) as f: return json.load(f)

IMPORT_META_FILE    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "import_meta.json")
PROCESSED_ORDERS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "processed_orders.json")

def load_import_meta():
    if os.path.exists(IMPORT_META_FILE):
        try:
            with open(IMPORT_META_FILE) as f: return json.load(f)
        except: pass
    return {}

def save_import_meta(meta: dict):
    with open(IMPORT_META_FILE, "w") as f: json.dump(meta, f, indent=2)

def load_processed_orders():
    """Returns {client: set(order_keys)} where order_key = 'ORDERID|DATE|SCRIP|SIDE'"""
    if os.path.exists(PROCESSED_ORDERS_FILE):
        try:
            with open(PROCESSED_ORDERS_FILE) as f:
                raw = json.load(f)
            return {c: set(v) for c, v in raw.items()}
        except: pass
    return {}

def save_processed_orders(po: dict):
    with open(PROCESSED_ORDERS_FILE, "w") as f:
        json.dump({c: list(v) for c, v in po.items()}, f, indent=2)

def make_order_key(order_no, trade_date, scrip, side):
    """Unique fingerprint for one broker order line."""
    date_str = str(trade_date)[:10] if trade_date else ''
    no_str   = str(order_no).strip().lstrip("'")
    return f"{no_str}|{date_str}|{str(scrip).strip().upper()}|{str(side).strip().upper()}"

def save(t):
    with open(DATA_FILE,"w") as f: json.dump(t,f,indent=2)
    # auto-update import meta: find latest trade date per client
    try:
        meta = load_import_meta()
        df_m = pd.DataFrame(t)
        if not df_m.empty:
            for c, grp in df_m.groupby('client'):
                raw = []
                for col in ['entry_date', 'exit_date']:
                    if col in grp.columns:
                        raw.extend(grp[col].dropna().tolist())
                dates = pd.to_datetime(raw, errors='coerce').dropna()
                if not dates.empty:
                    latest = dates.max().strftime('%d %b %Y')
                    # only overwrite if not already set by manual upload
                    if c not in meta:
                        meta[c] = f"Auto-detected: {latest}"
            save_import_meta(meta)
    except Exception:
        pass

CHARGE_COLS = ["buy_brokerage","buy_stt","buy_gst","buy_stamp","buy_txn",
               "sell_brokerage","sell_stt","sell_gst","sell_stamp","sell_txn"]

# ── Ledger (payin / payout) ───────────────────────────────────────────────────
def load_ledger() -> dict:
    if not os.path.exists(LEDGER_FILE): return {}
    with open(LEDGER_FILE) as f: return json.load(f)

def save_ledger(data: dict):
    with open(LEDGER_FILE, "w") as f: json.dump(data, f, indent=2)

def _parse_amount(val) -> float:
    """Parse Indian-formatted amount string: '1,00,000.00' → 100000.0"""
    try:
        return float(str(val).replace(",", "").strip())
    except Exception:
        return 0.0

def parse_ledger_csv(file_bytes, client: str, ledger_type: str) -> list:
    """Parse MO payin/payout CSV → list of entry dicts."""
    import io as _io
    df = pd.read_csv(_io.BytesIO(file_bytes))
    df.columns = [c.strip() for c in df.columns]
    entries = []
    for _, row in df.iterrows():
        vtype = str(row.get("VOUCHER TYPE", "")).strip().upper()
        if vtype not in ("PAYIN", "PAYOUT"): continue
        debit  = _parse_amount(row.get("DEBIT",  0))
        credit = _parse_amount(row.get("CREDIT", 0))
        amount = credit if vtype == "PAYIN" else -debit
        if amount == 0: continue
        raw_date = str(row.get("VOUCHER DATE", "")).strip()
        try:
            dt = pd.to_datetime(raw_date, dayfirst=True).strftime("%Y-%m-%d")
        except Exception:
            continue
        entries.append({
            "date":      dt,
            "type":      vtype,
            "amount":    round(amount, 2),
            "ledger":    ledger_type,   # "MTF" or "Delivery"
            "narration": str(row.get("NARRATION", "")).strip(),
        })
    return entries

def ledger_summary(client: str) -> dict:
    """Returns {payin, payout, interest, net} for a client."""
    entries = load_ledger().get(client, [])
    payin    = sum(e["amount"] for e in entries if e["type"] == "PAYIN")
    payout   = sum(-e["amount"] for e in entries if e["type"] == "PAYOUT")
    interest = sum(-e["amount"] for e in entries if e["type"] == "INTEREST")
    return {"payin": payin, "payout": payout, "interest": interest, "net": payin - payout}

def as_df():
    trades = load()
    if not trades: return pd.DataFrame()
    df = pd.DataFrame(trades)
    df["entry_date"] = pd.to_datetime(df["entry_date"], errors="coerce")
    df["exit_date"]  = pd.to_datetime(df["exit_date"],  errors="coerce")
    # Sanitize impossible trades: exit before entry → treat as open (never show as closed)
    bad_mask = df["exit_date"].notna() & (df["exit_date"] < df["entry_date"])
    if bad_mask.any():
        df.loc[bad_mask, "exit_date"]     = pd.NaT
        df.loc[bad_mask, "sell_qty"]      = 0
        df.loc[bad_mask, "sell_price"]    = 0
        df.loc[bad_mask, "sell_net_rate"] = 0
    num_cols = ["buy_qty","buy_price","buy_net_rate","sell_qty","sell_price","sell_net_rate"] + CHARGE_COLS
    for c in num_cols:
        if c not in df.columns: df[c] = 0.0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    # qty should always be whole numbers
    for c in ["buy_qty","sell_qty"]:
        if c in df.columns: df[c] = df[c].astype(int)
    df["position"]    = df.apply(lambda r: "Closed" if r["sell_qty"]>=r["buy_qty"] and r["sell_qty"]>0 else "Open", axis=1)
    df["invested"]    = df["buy_qty"]*df["buy_price"]
    today             = pd.Timestamp.now()
    df["pnl"]         = df.apply(lambda r: (r["sell_price"]-r["buy_price"])*r["sell_qty"] if r["position"]=="Closed" else 0, axis=1)
    df["net_pnl"]     = df.apply(lambda r: (r["sell_net_rate"]-r["buy_net_rate"])*r["sell_qty"] if r["position"]=="Closed" else 0, axis=1)
    df["return_pct"]  = df.apply(lambda r: (r["sell_price"]-r["buy_price"])/r["buy_price"]*100 if r["position"]=="Closed" and r["buy_price"] else 0, axis=1)
    df["net_ret_pct"] = df.apply(lambda r: (r["sell_net_rate"]-r["buy_net_rate"])/r["buy_net_rate"]*100 if r["position"]=="Closed" and r["buy_net_rate"] else 0, axis=1)
    def calc_days(r):
        if not pd.notna(r["entry_date"]): return 0
        end = r["exit_date"] if (r["position"]=="Closed" and pd.notna(r["exit_date"])) else today
        return max(0,(end-r["entry_date"]).days)
    df["days_held"]   = df.apply(calc_days, axis=1)
    # total charges per trade
    df["total_charges"] = df[CHARGE_COLS].sum(axis=1)
    # intraday flag
    df["intraday"] = df.apply(
        lambda r: r["position"]=="Closed" and pd.notna(r["exit_date"]) and r["entry_date"].date()==r["exit_date"].date(), axis=1)
    return df

def fmt(v):
    try:
        v=float(v); s="-" if v<0 else ""
        return f"{s}₹{abs(v):,.0f}"
    except: return "—"

def color_pnl(val):
    """For st.dataframe styling"""
    try:
        v = float(str(val).replace("₹","").replace(",","").replace("+",""))
        if v > 0: return "color:#34d399; font-weight:700"
        if v < 0: return "color:#f87171; font-weight:700"
    except: pass
    return ""

def styled(df_show, pnl_cols):
    try:
        try:
            return df_show.style.map(color_pnl, subset=pnl_cols)      # pandas >= 2.1
        except AttributeError:
            return df_show.style.applymap(color_pnl, subset=pnl_cols)  # pandas < 2.1
    except Exception:
        return df_show  # fallback: no styling if jinja2 missing

# ── NSE Symbol Map (script name → Yahoo Finance ticker) ──────────────────────
SYMBOL_MAP = {
    "ALKYL AMINES":               "ALKYLAMINE",
    "BAJAJ HOUSING FINANCE LIMITED":"BAJAJHFL",
    "BALKRISHNA IND":             "BALKRISIND",
    "BHARAT COKING COAL":         "BHARATCOAL",
    "BOMBAY BURMAH TRADING COR":  "BBTC",
    "CAMS":                       "CAMS",
    "CENTURY PLYBOARDS":          "CENTURYPLY",
    "CHOLAMANDALAM FIN HOL LTD":  "CHOLAHLDNG",
    "COFORGE":                    "COFORGE",
    "DATA PATTERNS":              "DATAPATTNS",
    "DEVYANI INTERNATIONAL":      "DEVYANI",
    "DIXON TECHNOLOGIES":         "DIXON",
    "DOMS INDUSTRIES":            "DOMS",
    "EQUITAS SMALL FIN BANK":     "EQUITASBNK",
    "EXIDE INDUSTRIES":           "EXIDEIND",
    "FORTIS HEALTHCARE LTD":      "FORTIS",
    "GARWARE TECH FIBRES":        "GARFIBRES",
    "GMR AIRPORTS":               "GMRAIRPORT",
    "GODREJ INDUSTRIES":          "GODREJIND",
    "GODREJ PROPERTIES":          "GODREJPROP",
    "HDFC BANK LTD":              "HDFCBANK",
    "HIMADRI SPECIALITY CHEM":    "HSCL",
    "HINDUSTAN COPPER LTD":       "HINDCOPPER",
    "HINDUSTAN COPPER LTD.":      "HINDCOPPER",
    "HINDUSTAN ZINC LTD.":        "HINDZINC",
    "IDBI BANK LTD.":             "IDBI",
    "INDIAN ENERGY EXCHANGE":     "IEX",
    "INDIAN ENERGY EXCHANGE LIMITED": "IEX",
    "INDIAN HOTELS":              "INDHOTEL",
    "INDOCO REMEDIES":            "INDOCO",
    "INFOSYS LIMITED":            "INFY",
    "INTELLECT DESIGN ARENA":     "INTELLECT",
    "INTERGLOBE AVIATION":        "INDIGO",
    "ITC LTD":                    "ITC",
    "JBM AUTO":                   "JBMA",
    "JM FINANCIAL":               "JMFINANCIL",
    "JM FINANCIAL LIMITED":       "JMFINANCIL",
    "KALPATARU":                  "KALPATARU",
    "KALPATARU PROJECTS":         "KPIL",
    "KPIT TECHNOLOGIES LIMITED":  "KPITTECH",
    "L&T TECHNOLOGY SERVICES":    "LTTS",
    "LIFE INSURANCE CORPORATION OF": "LICI",
    "LUPIN":                      "LUPIN",
    "MIC ELECTRONICS":            "MICEL",
    "MIRAE ENERGY ETF":           "MAFANG",
    "MRPL":                       "MRPL",
    "NATCO PHARMA":               "NATCOPHARM",
    "NELCO":                      "NELCO",
    "NETWEB TECHNOLOGIES":        "NETWEB",
    "NETWEB TECHNOLOGIES INDIA LIMI": "NETWEB",
    "NEWGEN SOFTWARE TECH LTD":   "NEWGEN",
    "NIPPON NETFSILVER":          "SILVERBEES",
    "OBEROI REALTY":              "OBEROIRLTY",
    "ORACLE FINANCIAL SERVICES":  "OFSS",
    "ORACLE FINANCIAL SERVICES SOFT": "OFSS",
    "PG ELECTROPLAST":            "PGEL",
    "PI INDUSTRIES":              "PIIND",
    "PIRAMAL PHARMA":             "PPLPHARMA",
    "POONAWALLA FINCORP":         "POONAWALLA",
    "REC":                        "RECLTD",
    "REDINGTON":                  "REDINGTON",
    "RELIANCE INDUSTRIES LTD":    "RELIANCE",
    "SAGILITY LIMITED":           "SAGILITY",
    "SAMVARDHANA MOTHERSON":      "MOTHERSON",
    "SHAILY ENGINEERING":         "SHAILY",
    "SOUTH INDIAN BANK":          "SOUTHBANK",
    "SPANDANA SPHOORTY":          "SPANDANA",
    "SUZLON ENERGY":              "SUZLON",
    "TANLA PLATFORMS":            "TANLA",
    "TEJAS NETWORKS":             "TEJASNET",
    "VST INDUSTRIES":             "VSTIND",
    "WHIRLPOOL INDIA":            "WHIRLPOOL",
    "WIPRO":                      "WIPRO",
    "WOCKHARDT":                  "WOCKPHARMA",
    "ZEN TECHNOLOGIES":           "ZENTEC",
    "ZENSAR TECHNOLOGIES":        "ZENSARTECH",
    "ANGEL ONE":                  "ANGELONE",
    "BANK OF INDIA":              "BANKINDIA",
    "BLUE JET HEALTHCARE":        "BLUEJET",
    "GRAPHITE INDIA":             "GRAPHITE",
    "GRINDWELL NORTON":           "GRINDWELL",
    "IIFL FINANCE":               "IIFL",
    "IFCI":                       "IFCI",
    "NMDC":                       "NMDC",
    "NTPC":                       "NTPC",
    "PARADEEP PHOSPHATES":        "PARADEEP",
    "PERSISTENT SYSTEMS":         "PERSISTENT",
    "VARUN BEVERAGES":            "VBL",
    "MAZAGON DOCK":               "MAZDOCK",
    "ALEMBIC PHARMA":             "APLLTD",
    "ASHAPURA MINECHEM":          "ASHAPURMIN",
    "GUJARAT NARMADA FERT":       "GNFC",
    "EMMVEE PHOTOVOLTAIC":        "EMMVEE",
    "MINDA CORPORATION":          "MINDACORP",
    "PHOENIX MILLS":              "PHOENIXLTD",
    "EICHER MOTORS":              "EICHERMOT",
    "BANK OF MAHARASHTRA":        "MAHABANK",
    "ZERODHAAMC - LIQUIDCASE":    "LIQUIDCASE",
    "INOX WIND":                  "INOXWIND",
    "MOSCHIP TECHNOLOGIES":       "MOSCHIP",
    "ZYDUS LIFESCIENCES":         "ZYDUSLIFE",
    "NIPPON LIFE INDIA ASSET MANAGE": "NAM-INDIA",
    "CANARA ROBECO ASSET MANAGEMENT": "CRAMC",
    # additional mappings from reference sheet
    "WIPRO JULFUT":               "WIPRO",
    "MMTC":                       "MMTC",
    "NATCOPHARM":                 "NATCOPHARM",
    "SPANDANA SPHOORTY FINANCIAL": "SPANDANA",
    "GRINDWELL NORTON LTD":       "GRINDWELL",
    "WOCKHARDT LTD":              "WOCKPHARMA",
    "NETWEB TECHNOLOGIES INDIA":  "NETWEB",
    "INDHOTEL":                   "INDHOTEL",
    "TEJAS NETWORKS LTD":         "TEJASNET",
    "REDINGTON INDIA":            "REDINGTON",
    "SHAILY ENGINEERING PLASTICS": "SHAILY",
    "GMR AIRPORTS INFRASTRUCTURE": "GMRAIRPORT",
    "GMR AIRPORTS LTD":           "GMRAIRPORT",
    "INOX WIND LTD":              "INOXWIND",
    "INOXWIND":                   "INOXWIND",
    "EMMVEE":                     "EMMVEE",
    "MIC ELECTRONICS LTD":        "MICEL",
    "BHARAT COKING COAL LTD":     "BHARATCOAL",
    "KPIL":                       "KPIL",
    "LUPIN LTD":                  "LUPIN",
    "GODREJ PROPERTIES LTD":      "GODREJPROP",
    "BAJAJ HOUSING FINANCE":      "BAJAJHFL",
    "LIFE INSURANCE CORP":        "LICI",
    "INTERGLOBE AVIATION LTD":    "INDIGO",
    "ITC LIMITED":                "ITC",
    "HDFC BANK":                  "HDFCBANK",
    "INFOSYS":                    "INFY",
    # confirmed tickers
    "BHARAT ELECTRONICS":         "BEL",
    "BOMBAY BURMAH TRADING":      "BBTC",
    "CANARA BANK":                "CANBK",
    "CHOLAMANDALAM":              "CHOLAFIN",
    "HINDUSTAN COPPER":           "HINDCOPPER",
    "IDBI BANK":                  "IDBI",
    "KPIT TECHNOLOGIES":          "KPITTECH",
    "LIC":                        "LICI",
    "MSTC":                       "MSTCLTD",
    "NEWGEN SOFTWARE":            "NEWGEN",
    "NIPPON LIFE AMC":            "NAM-INDIA",
    "RELIANCE INDUSTRIES":        "RELIANCE",
    "TRANSPORT CORP OF INDIA":    "TCI",
    "MIRAEAMC MAGOLDETF":         "GOLDETF",
    "NIPPON ETF IT":              "ITBEES",
    "NIPPON ETF LIQUID BEES":     "LIQUIDBEES",
    # additional scripts from trades data
    "ALOK INDUSTRIES LIMITED":    "ALOKINDS",
    "ARVIND FASHIONS LIMITED":    "ARVINDFAST",
    "AVENUE SUPERMARTS LIMITED":  "DMART",
    "BIOCON LIMITED.":            "BIOCON",
    "BSE LIMITED":                "BSE",
    "DELHIVERY":                  "DELHIVERY",
    "ELGI EQUIPMENTS LTD":        "ELGIEQUIP",
    "FORTIS HEALTHCARE":          "FORTIS",
    "HINDUSTAN ZINC":             "HINDZINC",
    "IDFC FIRST BANK":            "IDFCFIRSTB",
    "IRB INFRA DEV LTD.":         "IRBINFRA",
    "ITC":                        "ITC",
    "JSW ENERGY":                 "JSWENERGY",
    "JUPITER WAGONS LIMITED":     "JUPITERWAG",
    "K.P.R.MILL LIMITED":         "KPRMILL",
    "KAYNES TECHNOLOGY IND LTD":  "KAYNES",
    "LARSEN & TOUBRO LTD.":       "LT",
    "LAURUS LABS":                "LAURUSLABS",
    "M&M FINANCE":                "M&MFIN",
    "MAX HEALTHCARE INS LTD":     "MAXHEALTH",
    "NIIT LIMITED":               "NIITLTD",
    "POWER FIN CORP LTD.":        "PFC",
    "PREMIER ENERGIES LIMITED":   "PREMIERENE",
    "TCS":                        "TCS",
    "CEMINDIA PROJECTS LIMITED":  "CEMPRO",
    "JAIN RESOURCE RECYCLING LIMITE": "JAINREC",
}

# ── Ticker overrides (user-editable, stored in ticker_overrides.json) ────────
TICKER_OVERRIDES_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ticker_overrides.json")

def load_ticker_overrides():
    if os.path.exists(TICKER_OVERRIDES_FILE):
        try:
            with open(TICKER_OVERRIDES_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_ticker_overrides(overrides: dict):
    with open(TICKER_OVERRIDES_FILE, "w") as f:
        json.dump(overrides, f, indent=2)

def script_to_ticker(script):
    """Map a normalised script name → NSE Yahoo Finance ticker (adds .NS).
    User overrides take priority over SYMBOL_MAP."""
    overrides = load_ticker_overrides()
    if script in overrides and overrides[script]:
        sym = overrides[script].strip().upper().replace(".NS","")
    else:
        sym = SYMBOL_MAP.get(script, script.replace(" ","").replace("&","").replace(".",""))
    return sym + ".NS"

# ── CMP via Yahoo Finance (single batch request — fast) ──────────────────────
@st.cache_data(ttl=180)
def fetch_cmp(scripts: tuple):
    """Fetch live CMP using yfinance (reliable) with Yahoo Finance batch as fallback."""
    import yfinance as yf
    results  = {s: None for s in scripts}
    overrides = load_ticker_overrides()

    def get_ticker_sym(script):
        if script in overrides and overrides[script]:
            return overrides[script].strip().upper().replace(".NS","")
        return SYMBOL_MAP.get(script, script.replace(" ","").replace("&","").replace(".","").upper())

    # build ticker → script map
    t2s = {}
    for s in scripts:
        t = get_ticker_sym(s) + ".NS"
        t2s[t] = s

    # ── yfinance batch download (most reliable) ────────────────────────────────
    try:
        tickers_str = " ".join(t2s.keys())
        data = yf.download(tickers_str, period="1d", interval="1d",
                           progress=False, auto_adjust=True, threads=True)
        if not data.empty:
            close = data["Close"] if "Close" in data else data.get("close", pd.DataFrame())
            if isinstance(close, pd.Series):
                # single ticker returned as Series
                ticker = list(t2s.keys())[0]
                val    = close.dropna().iloc[-1] if not close.dropna().empty else None
                if val:
                    results[t2s[ticker]] = round(float(val), 2)
            else:
                for ticker, script in t2s.items():
                    if ticker in close.columns:
                        col = close[ticker].dropna()
                        if not col.empty:
                            results[script] = round(float(col.iloc[-1]), 2)
    except Exception:
        pass

    # ── fallback: yfinance fast_info per missing script ───────────────────────
    still_missing = [s for s in scripts if not results[s]]
    for script in still_missing:
        try:
            t   = yf.Ticker(get_ticker_sym(script) + ".NS")
            val = getattr(t.fast_info, "last_price", None) or getattr(t.fast_info, "regular_market_price", None)
            if val:
                results[script] = round(float(val), 2)
        except Exception:
            pass

    return results

def to_excel(df_dict):
    """Export one or more DataFrames to a formatted Excel workbook."""
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="1E2235")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL    = PatternFill("solid", fgColor="F4F5F9")
    GREEN_FILL  = PatternFill("solid", fgColor="D4EDDA")
    RED_FILL    = PatternFill("solid", fgColor="F8D7DA")
    GREEN_FONT  = Font(bold=True, color="155724")
    RED_FONT    = Font(bold=True, color="721C24")
    BORDER_SIDE = Side(style="thin", color="D0D3DC")
    THIN_BORDER = Border(bottom=BORDER_SIDE)
    RUPEE_FMT   = '₹#,##0.00'
    PCT_FMT     = '0.00"%"'

    # columns that should get currency formatting
    CURRENCY_COLS = {
        'buy price','sell price','avg buy price','buy','sell','invested','invested (₹)',
        'gross p&l','net p&l','total p&l','unrealized p&l','booked net p&l','booked gross p&l',
        'capital deployed','total charges','charges','buy charges','sell charges',
        'total invested','total invested (₹)','best trade','worst trade','net amount',
        'avg sell','avg buy',
    }
    PNL_COLS = {
        'gross p&l','net p&l','total p&l','unrealized p&l','booked net p&l',
        'booked gross p&l','notional p&l','best trade','worst trade',
    }
    PCT_COLS = {'% return','return %','net%','win rate %'}

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        for sheet_name, df_ in df_dict.items():
            if df_ is None or (hasattr(df_, 'empty') and df_.empty):
                continue
            df_ = df_.copy()
            # drop internal/noisy columns
            drop = [c for c in df_.columns if c.lower() in
                    {'id','type','notes','buy_brokerage','buy_stt','buy_gst','buy_stamp',
                     'buy_txn_chrg','buy_txn','sell_brokerage','sell_stt','sell_gst',
                     'sell_stamp','sell_txn_chrg','sell_txn','intraday',
                     'buy_net_rate','sell_net_rate','net_ret_pct'}]
            df_.drop(columns=drop, errors='ignore', inplace=True)
            # friendly column names
            rename = {
                'client':       'Client',
                'script':       'Script',
                'entry_date':   'Entry Date',
                'exit_date':    'Exit Date',
                'buy_qty':      'Qty',
                'buy_price':    'Buy Price',
                'sell_qty':     'Sell Qty',
                'sell_price':   'Sell Price',
                'pnl':          'Gross P&L',
                'net_pnl':      'Net P&L',
                'net_pnl_':     'Net P&L',
                'return_pct':   'Return %',
                'net_ret_pct':  'Net Return %',
                'days_held':    'Days Held',
                'invested':     'Invested (₹)',
                'buy_net_rate': 'Buy Net Rate',
                'sell_net_rate':'Sell Net Rate',
                'total_charges':'Charges',
                'position':     'Status',
            }
            df_.rename(columns=rename, inplace=True)
            # format client codes to names
            if 'Client' in df_.columns:
                df_['Client'] = df_['Client'].apply(lambda c: CLIENT_NAMES.get(c, c))
            # format dates
            for col in df_.columns:
                if 'date' in col.lower() or col in ('Entry Date','Exit Date'):
                    df_[col] = pd.to_datetime(df_[col], errors='coerce').dt.strftime('%d/%m/%Y').fillna('')

            df_.to_excel(w, sheet_name=sheet_name[:31], index=False)
            ws = w.sheets[sheet_name[:31]]

            ncols = len(df_.columns)
            nrows = len(df_)

            # header row styling
            for col_idx, col_name in enumerate(df_.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill   = HEADER_FILL
                cell.font   = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = THIN_BORDER

            # data rows
            col_names_lower = [c.lower() for c in df_.columns]
            for row_idx in range(2, nrows + 2):
                is_alt = (row_idx % 2 == 0)
                for col_idx, col_name in enumerate(col_names_lower, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # alternating row fill (skip if P&L will override)
                    if is_alt and col_name not in PNL_COLS:
                        cell.fill = ALT_FILL
                    cell.alignment = Alignment(horizontal='right' if col_name in CURRENCY_COLS | PCT_COLS else 'left',
                                               vertical='center')
                    # number formats
                    if col_name in CURRENCY_COLS:
                        cell.number_format = RUPEE_FMT
                    elif col_name in PCT_COLS:
                        cell.number_format = '0.00'

            # conditional green/red on P&L columns
            for col_idx, col_name in enumerate(col_names_lower, 1):
                if col_name in PNL_COLS and nrows > 0:
                    col_letter = get_column_letter(col_idx)
                    data_range = f"{col_letter}2:{col_letter}{nrows+1}"
                    ws.conditional_formatting.add(data_range,
                        CellIsRule(operator='greaterThan', formula=['0'],
                                   fill=GREEN_FILL, font=GREEN_FONT))
                    ws.conditional_formatting.add(data_range,
                        CellIsRule(operator='lessThan', formula=['0'],
                                   fill=RED_FILL, font=RED_FONT))

            # auto column widths
            for col_idx, col_name in enumerate(df_.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    len(str(col_name)),
                    df_[col_name].astype(str).str.len().max() if nrows > 0 else 0
                )
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 30)

            # freeze header row
            ws.freeze_panes = 'A2'
            # row height for header
            ws.row_dimensions[1].height = 30

    return buf.getvalue()

# ── CLIENT COLORS ─────────────────────────────────────────────────────────────
CLIENT_COLORS = {
    "RIMK1209": "#818cf8",   # indigo   – Sathyavrath
    "RIMK1220": "#fb923c",   # orange   – Kalpana
    "RIMK1238": "#a78bfa",   # violet   – Iranna
    "RIMK1248": "#34d399",   # emerald  – Udayakumar
    "RIMK1249": "#fbbf24",   # amber    – Sundareshwari
    "RIMK1252": "#f472b6",   # pink     – Savitha
}

# ── TOP NAVBAR ────────────────────────────────────────────────────────────────
_ALL_PAGES   = ["Overview","Insights","Positions","Stock Analysis","Brokerage","💰 Capital","Add / Close","Settings"]
_ADMIN_PAGES = {"Add / Close", "Settings", "💰 Capital"}
PAGES = _ALL_PAGES if _IS_ADMIN else [p for p in _ALL_PAGES if p not in _ADMIN_PAGES]
PAGE_ICONS = {"Overview":"📊","Insights":"🔍","Positions":"📋",
              "Stock Analysis":"📌","Brokerage":"💸","💰 Capital":"💰","Add / Close":"➕","Settings":"⚙️"}

if "page" not in st.session_state or st.session_state.page not in PAGES:
    st.session_state.page = "Overview"

# Brand + user info
st.markdown("""
<style>
.nav-brand-bar {
    background:#080a10; border-bottom:1px solid #1f2232;
    padding:0 24px; display:flex; align-items:center; justify-content:space-between;
    height:50px; margin:0 -2.5rem;
}
.nav-brand-title { font-size:1rem; font-weight:800; color:#fff; letter-spacing:-.01em; }
.nav-brand-title em { color:#818cf8; font-style:normal; }
.nav-user { font-size:.78rem; color:#8890a8; }
</style>""", unsafe_allow_html=True)
st.markdown(f"""
<div class="nav-brand-bar">
  <div class="nav-brand-title">📈 Client <em>Tracker MOFSL</em></div>
  <div class="nav-user">👤 {st.session_state.auth_user} &nbsp;·&nbsp; {"Admin" if _IS_ADMIN else "Viewer"}</div>
</div>
""", unsafe_allow_html=True)

# Nav rows + logout
nav_row1 = PAGES[:4]
nav_row2 = PAGES[4:]
for row in [nav_row1, nav_row2]:
    _extra = 1 if row is nav_row2 else 0
    cols = st.columns(len(row) + _extra)
    for i, p in enumerate(row):
        with cols[i]:
            is_active = st.session_state.page == p
            if st.button(f"{PAGE_ICONS[p]} {p}", key=f"nav_{p}",
                         use_container_width=True,
                         type="primary" if is_active else "secondary"):
                st.session_state.page = p
                st.rerun()
    if _extra:
        with cols[-1]:
            if st.button("🚪 Logout", key="nav_logout", use_container_width=True):
                st.session_state.auth_ok   = False
                st.session_state.auth_user = ""
                st.session_state.auth_role = ""
                st.rerun()

page = st.session_state.page
st.markdown("<hr style='margin:4px 0 12px 0;'>", unsafe_allow_html=True)

# ── CLIENT FILTER BAR ─────────────────────────────────────────────────────────
df_all = as_df()
if df_all.empty:
    st.info("No trades. Add via **Add / Close** or run `import_all.py`.")
    st.stop()

_closed_all = df_all[df_all["position"]=="Closed"]

def _pill_label(c):
    net = _closed_all[_closed_all["client"]==c]["net_pnl"].sum()
    pnl = f"₹{net/1000:.1f}k" if abs(net)>=1000 else f"₹{net:,.0f}"
    sign = "▲" if net >= 0 else "▼"
    return f"{CLIENT_NAMES[c]}  {sign}{pnl}"

sel_clients = st.pills(
    "Clients",
    options=CLIENTS,
    format_func=_pill_label,
    selection_mode="multi",
    default=CLIENTS,
    label_visibility="collapsed",
)

if not sel_clients:
    st.warning("Select at least one client.")
    st.stop()

st.markdown("<hr style='margin:6px 0 14px 0;'>", unsafe_allow_html=True)

df        = df_all[df_all["client"].isin(sel_clients)]
open_df   = df[df["position"]=="Open"]
closed_df = df[df["position"]=="Closed"]
intra_df  = df[df["intraday"]==True]
deliv_df  = closed_df[closed_df["intraday"]==False]

CHART_THEME = dict(
    plot_bgcolor  = "#0d0f1a",
    paper_bgcolor = "#0d0f1a",
    font          = dict(color="#9ca3af", size=12, family="Inter, sans-serif"),
    margin        = dict(t=36,b=24,l=8,r=8),
    xaxis         = dict(gridcolor="#1a1d2e", zerolinecolor="#1a1d2e",
                         linecolor="#1f2232", tickfont=dict(color="#6b7280",size=11)),
    yaxis         = dict(gridcolor="#1a1d2e", zerolinecolor="#1a1d2e",
                         linecolor="#1f2232", tickfont=dict(color="#6b7280",size=11)),
)
UP_COL  = "#34d399"   # emerald green
DN_COL  = "#f87171"   # soft red
ACC_COL = "#818cf8"   # indigo accent

def pcol(v): return UP_COL if v >= 0 else DN_COL

# ═══════════════════════════════════════════════════════════════════════════════
if page == "Overview":
# ═══════════════════════════════════════════════════════════════════════════════
    # ── Today's Activity Banner ──────────────────────────────────────────────
    from datetime import date as _date_cls
    _today_str = _date_cls.today().isoformat()
    _today_closed = closed_df[closed_df["exit_date"].dt.date == _date_cls.today()] if not closed_df.empty else pd.DataFrame()
    _today_open   = open_df[open_df["entry_date"].dt.date == _date_cls.today()]  if not open_df.empty  else pd.DataFrame()
    if not _today_closed.empty or not _today_open.empty:
        st.markdown('<div class="kpi-head">Today\'s Activity</div>', unsafe_allow_html=True)
        ta1, ta2, ta3, ta4 = st.columns(4)
        ta_pnl = _today_closed["net_pnl"].sum() if not _today_closed.empty else 0
        ta_trades = len(_today_closed) + len(_today_open)
        ta_clients = pd.concat([
            _today_closed["client"] if not _today_closed.empty else pd.Series(dtype=str),
            _today_open["client"]   if not _today_open.empty   else pd.Series(dtype=str)
        ]).nunique()
        ta_vol = ((_today_closed["sell_price"]*_today_closed["sell_qty"]).sum()
                  if not _today_closed.empty else 0)
        ta1.metric("Today's Trades",   str(ta_trades))
        ta2.metric("Clients Active",   str(ta_clients))
        ta3.metric("Today's Net P&L",  fmt(ta_pnl),  delta=f"{'▲' if ta_pnl>=0 else '▼'}")
        ta4.metric("Turnover Today",   fmt(ta_vol))
        st.markdown("")

    # ── Capital Snapshot ──────────────────────────────────────────────────────
    _led_data   = load_ledger()
    _any_ledger = any(bool(_led_data.get(c)) for c in CLIENTS)
    if _any_ledger:
        st.markdown('<div class="kpi-head">Capital Snapshot</div>', unsafe_allow_html=True)
        _tot_payin = _tot_payout = 0
        for _c in CLIENTS:
            _s = ledger_summary(_c)
            _tot_payin  += _s["payin"]
            _tot_payout += _s["payout"]
        _net_cap = _tot_payin - _tot_payout
        _deployed = open_df["invested"].sum() if not open_df.empty else 0
        _booked   = closed_df["net_pnl"].sum() if not closed_df.empty else 0
        _util_pct = _deployed / _net_cap * 100 if _net_cap > 0 else 0
        cs1, cs2, cs3, cs4 = st.columns(4)
        cs1.metric("Net Capital In",    f"₹{_net_cap/1e5:.2f}L")
        cs2.metric("Deployed (Open)",   f"₹{_deployed/1e5:.2f}L", delta=f"{_util_pct:.0f}% utilised")
        cs3.metric("Booked Net P&L",    fmt(_booked))
        cs4.metric("Overall ROI",       f"{_booked/_net_cap*100:+.2f}%" if _net_cap else "—")
        st.markdown("")

    gross   = closed_df["pnl"].sum()
    net     = closed_df["net_pnl"].sum()
    charges = closed_df["total_charges"].sum()
    cap     = open_df["invested"].sum()
    wins    = (closed_df["net_pnl"]>0).sum()
    losses  = (closed_df["net_pnl"]<0).sum()
    wr      = wins/len(closed_df)*100 if len(closed_df) else 0
    avg_w   = closed_df[closed_df["net_pnl"]>0]["net_ret_pct"].mean() if wins else 0
    avg_l   = closed_df[closed_df["net_pnl"]<0]["net_ret_pct"].mean() if losses else 0
    best    = closed_df["net_pnl"].max() if len(closed_df) else 0
    worst   = closed_df["net_pnl"].min() if len(closed_df) else 0

    # ── row 1: P&L KPIs ──
    st.markdown('<div class="kpi-head">Realised Performance</div>', unsafe_allow_html=True)
    k1,k2,k3,k4 = st.columns(4)
    k1.metric("Gross P&L",     fmt(gross),   delta=f"Net {fmt(net)}")
    k2.metric("Net P&L",       fmt(net),     delta=f"After ₹{charges:,.0f} charges")
    k3.metric("Charges Paid",  fmt(charges), delta=f"{charges/gross*100:.1f}% of gross" if gross else None)
    k4.metric("Win Rate",      f"{wr:.0f}%", delta=f"{wins}W / {losses}L")

    # ── row 2: Portfolio KPIs ──
    st.markdown('<div class="kpi-head">Portfolio Status</div>', unsafe_allow_html=True)
    k5,k6,k7,k8 = st.columns(4)
    k5.metric("Open Capital",   fmt(cap),         delta=f"{len(open_df)} positions")
    k6.metric("Avg Win",        f"+{avg_w:.1f}%")
    k7.metric("Avg Loss",       f"{avg_l:.1f}%")
    k8.metric("Best / Worst",   f"{fmt(best)} / {fmt(worst)}")

    # ── row 3: Unrealized P&L ──
    st.markdown('<div class="kpi-head">Unrealised P&L (Live)</div>', unsafe_allow_html=True)
    _ucol1, _ucol2 = st.columns([1, 5])
    _fetch_live = _ucol1.button("📡 Fetch Live CMP", key="ov_fetch_cmp")
    if _fetch_live or st.session_state.get("ov_cmp_cache"):
        if _fetch_live:
            with st.spinner("Fetching live prices from Yahoo Finance..."):
                _cmp = fetch_cmp(tuple(open_df["script"].unique()))
            st.session_state.ov_cmp_cache = _cmp
        else:
            _cmp = st.session_state.ov_cmp_cache
        _total_invested = open_df["invested"].sum()
        _total_current  = open_df.apply(
            lambda r: _cmp.get(r["script"], r["buy_price"]) * r["buy_qty"], axis=1).sum()
        _unreal = _total_current - _total_invested
        _unreal_pct = _unreal / _total_invested * 100 if _total_invested else 0
        _u1, _u2, _u3 = st.columns(3)
        _u1.metric("Invested Capital", fmt(_total_invested))
        _u2.metric("Current Value",    fmt(_total_current),
                   delta=f"{_unreal_pct:+.1f}%")
        _u3.metric("Unrealised P&L",   fmt(_unreal),
                   delta=f"{_unreal_pct:+.1f}%")
        # Per-script table
        _urows = []
        for _, _r in open_df.iterrows():
            _cmp_val = _cmp.get(_r["script"])
            if _cmp_val:
                _upnl = (_cmp_val - _r["buy_price"]) * _r["buy_qty"]
                _urows.append({
                    "Client":  CLIENT_NAMES.get(_r["client"], _r["client"]),
                    "Script":  _r["script"],
                    "Qty":     int(_r["buy_qty"]),
                    "Buy Avg": f"₹{_r['buy_price']:,.2f}",
                    "CMP":     f"₹{_cmp_val:,.2f}",
                    "Invested": fmt(_r["invested"]),
                    "Unreal P&L": fmt(_upnl),
                    "Return %":  f"{(_cmp_val-_r['buy_price'])/_r['buy_price']*100:+.1f}%",
                })
        if _urows:
            _udf = pd.DataFrame(_urows).sort_values("Unreal P&L", key=lambda s: s.str.replace("[₹,()% ]","",regex=True).str.replace("−","-").apply(lambda x: float(x) if x.lstrip("-").replace(".","").isdigit() else 0), ascending=False)
            st.dataframe(styled(_udf, ["Unreal P&L"]), use_container_width=True, hide_index=True, height=420)
    else:
        _ucol2.caption("Click to fetch live NSE prices and see unrealised P&L for all open positions.")

    st.markdown("---")

    # ── Charts row ──
    ch1, ch2 = st.columns([3, 2])

    with ch1:
        st.markdown('<div class="section-head">Monthly Net P&L</div>', unsafe_allow_html=True)
        if not closed_df.empty:
            mdf = closed_df.copy()
            mdf["month"] = mdf["exit_date"].dt.to_period("M").astype(str)
            monthly = mdf.groupby("month").agg(
                net_pnl=("net_pnl","sum"), trades=("pnl","count"),
                cum=("net_pnl","sum")
            ).reset_index()
            monthly["cum"] = monthly["net_pnl"].cumsum()
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=monthly["month"], y=monthly["net_pnl"],
                marker_color=[pcol(v) for v in monthly["net_pnl"]],
                marker_line_width=0,
                text=monthly["net_pnl"].apply(lambda x: f"₹{x/1000:.0f}k"),
                textposition="outside", name="Monthly Net P&L",
                hovertemplate="<b>%{x}</b><br>Net P&L: ₹%{y:,.0f}<br>Trades: %{customdata}<extra></extra>",
                customdata=monthly["trades"]
            ))
            fig.add_trace(go.Scatter(
                x=monthly["month"], y=monthly["cum"],
                mode="lines+markers", name="Cumulative",
                line=dict(color=ACC_COL, width=2.5),
                marker=dict(size=5), yaxis="y2",
                hovertemplate="Cumulative: ₹%{y:,.0f}<extra></extra>"
            ))
            fig.update_layout(**CHART_THEME, height=300, showlegend=True,
                              legend=dict(orientation="h",y=1.12,x=0),
                              yaxis2=dict(overlaying="y", side="right",
                                          gridcolor="#1a1f35", showgrid=False))
            st.plotly_chart(fig, use_container_width=True)

    with ch2:
        st.markdown('<div class="section-head">Net P&L by Client</div>', unsafe_allow_html=True)
        cdata = []
        for c in CLIENTS:
            if c not in sel_clients: continue
            cl = closed_df[closed_df["client"]==c]
            op = open_df[open_df["client"]==c]
            cdata.append({
                "Client": CLIENT_NAMES[c], "net": cl["net_pnl"].sum(),
                "color": CLIENT_COLORS[c]
            })
        cdf2 = pd.DataFrame(cdata).sort_values("net", ascending=True)
        if not cdf2.empty:
            fig2 = go.Figure(go.Bar(
                x=cdf2["net"], y=cdf2["Client"], orientation="h",
                marker_color=cdf2["color"], marker_line_width=0,
                text=cdf2["net"].apply(lambda x: f"₹{x/1000:.1f}k"),
                textposition="outside",
                hovertemplate="<b>%{y}</b>: ₹%{x:,.0f}<extra></extra>"
            ))
            fig2.update_layout(**CHART_THEME, height=300, showlegend=False,
                               xaxis_title=None, yaxis_title=None)
            st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")

    # ── Client cards ──
    st.markdown('<div class="section-head">Client Breakdown</div>', unsafe_allow_html=True)
    card_cols = st.columns(len(sel_clients))
    for i, c in enumerate(sel_clients):
        cl = closed_df[closed_df["client"]==c]
        op = open_df[open_df["client"]==c]
        it = intra_df[intra_df["client"]==c]
        net_ = cl["net_pnl"].sum()
        wr_  = (cl["net_pnl"]>0).sum()/len(cl)*100 if len(cl) else 0
        color = CLIENT_COLORS.get(c,"#888")
        pnl_cls = "up" if net_ >= 0 else "down"
        with card_cols[i]:
            st.markdown(f"""
            <div style="background:#13162a;border:1px solid #1e2235;border-top:3px solid {color};
                        border-radius:12px;padding:16px;text-align:center;">
              <div style="font-size:.72rem;color:{color};font-weight:700;letter-spacing:.08em;
                          text-transform:uppercase;">{c}</div>
              <div style="font-size:1.1rem;font-weight:800;margin:6px 0;">{CLIENT_NAMES[c]}</div>
              <div style="font-size:1.25rem;font-weight:800;" class="{pnl_cls}">{fmt(net_)}</div>
              <div style="font-size:.72rem;color:#8890a8;margin-top:4px;">Net P&L</div>
              <hr style="border-color:#1e2235;margin:10px 0;">
              <div style="display:flex;justify-content:space-between;font-size:.75rem;">
                <span>🟢 {len(op)} open</span>
                <span>✅ {len(cl)} closed</span>
                <span>⚡ {len(it)} intra</span>
              </div>
              <div style="font-size:.75rem;margin-top:6px;color:#8890a8;">
                Win rate: <b style="color:#fff;">{wr_:.0f}%</b> &nbsp;·&nbsp;
                Capital: <b style="color:#fff;">{fmt(op["invested"].sum())}</b>
              </div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    sdf_rows = []
    for c in sel_clients:
        cl = closed_df[closed_df["client"]==c]; op = open_df[open_df["client"]==c]
        it = intra_df[intra_df["client"]==c];   w  = (cl["net_pnl"]>0).sum()
        g_ = cl["pnl"].sum(); n_ = cl["net_pnl"].sum()
        sdf_rows.append({
            "Client": f"{CLIENT_NAMES[c]} ({c})", "Open": len(op), "Closed": len(cl),
            "Intraday": len(it), "Capital": fmt(op["invested"].sum()),
            "Gross P&L": fmt(g_), "Net P&L": fmt(n_),
            "Charges": fmt(g_-n_),
            "Win Rate": f"{w}/{len(cl)} ({w/len(cl)*100:.0f}%)" if len(cl) else "—",
            "Avg Hold": f"{cl['days_held'].mean():.0f}d" if len(cl) else "—",
        })
    sdf = pd.DataFrame(sdf_rows)
    st.dataframe(styled(sdf,["Gross P&L","Net P&L"]), use_container_width=True, hide_index=True)

    xls = to_excel({
        "P&L Summary":    sdf,
        "Open Positions": open_df.drop(columns=CHARGE_COLS, errors="ignore"),
        "Closed Trades":  closed_df.drop(columns=CHARGE_COLS, errors="ignore"),
    })
    st.download_button("📥 Download Portfolio Report (Excel)", xls, "portfolio_report.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Insights":
# ═══════════════════════════════════════════════════════════════════════════════
    st.markdown("## Insights")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown('<div class="section-head">Top 5 Winners (Net P&L)</div>', unsafe_allow_html=True)
        if not closed_df.empty:
            top = closed_df.nlargest(5,"net_pnl")[["client","script","buy_price","sell_price","pnl","net_pnl","return_pct","days_held"]].copy()
            top.columns = ["Client","Script","Buy","Sell","Gross P&L","Net P&L","Return %","Days"]
            top["Buy"]     = top["Buy"].apply(lambda x: f"₹{x:,.2f}")
            top["Sell"]    = top["Sell"].apply(lambda x: f"₹{x:,.2f}")
            top["Gross P&L"] = top["Gross P&L"].apply(fmt)
            top["Net P&L"] = top["Net P&L"].apply(fmt)
            top["Return %"] = top["Return %"].apply(lambda x: f"+{x:.1f}%")
            st.dataframe(styled(top,["Gross P&L","Net P&L"]), use_container_width=True, hide_index=True)

    with col2:
        st.markdown('<div class="section-head">Top 5 Losers (Net P&L)</div>', unsafe_allow_html=True)
        if not closed_df.empty:
            bot = closed_df.nsmallest(5,"net_pnl")[["client","script","buy_price","sell_price","pnl","net_pnl","return_pct","days_held"]].copy()
            bot.columns = ["Client","Script","Buy","Sell","Gross P&L","Net P&L","Return %","Days"]
            bot["Buy"]     = bot["Buy"].apply(lambda x: f"₹{x:,.2f}")
            bot["Sell"]    = bot["Sell"].apply(lambda x: f"₹{x:,.2f}")
            bot["Gross P&L"] = bot["Gross P&L"].apply(fmt)
            bot["Net P&L"] = bot["Net P&L"].apply(fmt)
            bot["Return %"] = bot["Return %"].apply(lambda x: f"{x:.1f}%")
            st.dataframe(styled(bot,["Gross P&L","Net P&L"]), use_container_width=True, hide_index=True)

    col3, col4 = st.columns(2)
    with col3:
        st.markdown('<div class="section-head">Return % Distribution</div>', unsafe_allow_html=True)
        if not closed_df.empty:
            fig = px.histogram(closed_df, x="net_ret_pct", nbins=25,
                               color_discrete_sequence=[ACC_COL], height=230)
            fig.add_vline(x=0, line_color=DN_COL, line_dash="dash", line_width=1.5)
            fig.update_layout(**CHART_THEME, xaxis_title="Net Return %", yaxis_title="# Trades")
            st.plotly_chart(fig, use_container_width=True)
    with col4:
        st.markdown('<div class="section-head">Holding Period Distribution</div>', unsafe_allow_html=True)
        if not closed_df.empty:
            fig2 = px.histogram(closed_df[closed_df["intraday"]==False], x="days_held", nbins=15,
                                color_discrete_sequence=["#7c4dff"], height=230)
            fig2.update_layout(**CHART_THEME, xaxis_title="Days Held", yaxis_title="# Trades")
            st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")
    st.markdown('<div class="section-head">P&L by Script (sorted by Net P&L)</div>', unsafe_allow_html=True)
    if not closed_df.empty:
        sp = closed_df.groupby("script").agg(
            trades=("pnl","count"), gross=("pnl","sum"), net=("net_pnl","sum"),
            charges=("total_charges","sum"), avg_ret=("net_ret_pct","mean"),
            wins=("net_pnl", lambda x:(x>0).sum())
        ).reset_index().sort_values("net",ascending=False)
        sp["win_rate"] = sp.apply(lambda r: f"{r['wins']:.0f}/{r['trades']:.0f} ({r['wins']/r['trades']*100:.0f}%)", axis=1)
        sp_show = sp[["script","trades","gross","net","charges","avg_ret","win_rate"]].copy()
        sp_show["gross"]   = sp_show["gross"].apply(fmt)
        sp_show["net"]     = sp_show["net"].apply(fmt)
        sp_show["charges"] = sp_show["charges"].apply(fmt)
        sp_show["avg_ret"] = sp_show["avg_ret"].apply(lambda x: f"{x:.1f}%")
        sp_show.columns    = ["Script","Trades","Gross P&L","Net P&L","Charges","Avg Net Return","Win Rate"]
        st.dataframe(styled(sp_show,["Gross P&L","Net P&L"]), use_container_width=True, hide_index=True, height=400)

    st.markdown("---")
    c1,c2,c3 = st.columns(3)
    with c1:
        st.markdown('<div class="section-head">Stale Positions (>30 days)</div>', unsafe_allow_html=True)
        stale = open_df[open_df["days_held"]>=30].sort_values("days_held",ascending=False)
        if stale.empty: st.success("None — clean book!")
        else:
            s = stale[["client","script","buy_price","buy_qty","invested","days_held"]].copy()
            s.columns=["Client","Script","Buy ₹","Qty","Invested","Days"]
            s["Invested"]=s["Invested"].apply(fmt)
            st.dataframe(s, use_container_width=True, hide_index=True, height=250)
    with c2:
        st.markdown('<div class="section-head">Largest Open Positions</div>', unsafe_allow_html=True)
        big = open_df.nlargest(8,"invested")[["client","script","buy_price","buy_qty","invested","days_held"]].copy()
        big.columns=["Client","Script","Buy ₹","Qty","Invested","Days"]
        big["Invested"]=big["Invested"].apply(fmt)
        st.dataframe(big, use_container_width=True, hide_index=True, height=280)
    with c3:
        st.markdown('<div class="section-head">Live CMP (Yahoo Finance)</div>', unsafe_allow_html=True)
        if st.button("Fetch CMP for open positions"):
            scripts = tuple(open_df["script"].unique())
            with st.spinner("Fetching..."):
                cmp_data = fetch_cmp(scripts)
            fetched = {k:v for k,v in cmp_data.items() if v}
            if fetched:
                rows_cmp = []
                for script, cmp in fetched.items():
                    sob = open_df[open_df["script"]==script]
                    avg_buy = (sob["buy_price"]*sob["buy_qty"]).sum()/sob["buy_qty"].sum()
                    notional = (cmp - avg_buy)*sob["buy_qty"].sum()
                    rows_cmp.append({
                        "Script": script,
                        "Ticker": script_to_ticker(script),
                        "CMP": f"₹{cmp:,.2f}",
                        "Avg Buy": f"₹{avg_buy:,.2f}",
                        "Notional P&L": fmt(notional),
                        "Return %": f"{(cmp-avg_buy)/avg_buy*100:+.1f}%",
                    })
                cmp_df = pd.DataFrame(rows_cmp)
                st.dataframe(styled(cmp_df,["Notional P&L"]), use_container_width=True, hide_index=True, height=280)
                missed = [s for s in scripts if not cmp_data.get(s)]
                if missed: st.caption(f"Could not fetch: {', '.join(missed)}")
            else:
                st.warning("Could not fetch any prices. Check internet / yfinance.")


# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Positions":
# ═══════════════════════════════════════════════════════════════════════════════
    # ── top-level filters ──
    fc1, fc2, fc3, fc4 = st.columns([2,2,2,2])
    search  = fc1.text_input("🔍 Search script")
    f_start = fc2.date_input("From",  value=date(2025,1,1))
    f_end   = fc3.date_input("To",    value=date.today())
    f_type  = fc4.selectbox("Trade type", ["All","Delivery","Intraday"])

    tab1, tab2, tab3 = st.tabs([f"🟢 Open ({len(open_df)})",
                                 f"🔴 Closed ({len(deliv_df)}) Delivery",
                                 f"⚡ Intraday ({len(intra_df)})"])

    def filter_df(d):
        if search: d = d[d["script"].str.contains(search.upper(), na=False)]
        if "entry_date" in d.columns:
            d = d[(d["entry_date"].dt.date >= f_start) & (d["entry_date"].dt.date <= f_end)]
        return d

    with tab1:
        odf = filter_df(open_df.copy())

        # ── Excel-style column filters ──
        xf1, xf2, xf3, xf4, xf5 = st.columns([2,2,2,2,1])
        f_client_open = xf1.multiselect("Client",
            options=sorted(odf["client"].unique()),
            format_func=lambda c: CLIENT_NAMES.get(c,c),
            default=list(odf["client"].unique()), key="f_cli_open",
            label_visibility="visible")
        f_script_open = xf2.text_input("Script contains", key="f_scr_open")
        f_minqty      = xf3.number_input("Min Qty", min_value=0, value=0, step=1, key="f_qty_open")
        f_maxdays     = xf4.number_input("Max Days Held", min_value=0, value=9999, step=1, key="f_days_open")
        fetch_open_cmp = xf5.button("📡 CMP", key="open_cmp_btn", use_container_width=True)

        if f_client_open: odf = odf[odf["client"].isin(f_client_open)]
        if f_script_open: odf = odf[odf["script"].str.contains(f_script_open.upper(), na=False)]
        if f_minqty > 0:  odf = odf[odf["buy_qty"] >= f_minqty]
        if f_maxdays < 9999: odf = odf[odf["days_held"] <= f_maxdays]

        cmp_open = {}
        if fetch_open_cmp:
            with st.spinner("Fetching CMP for all open scripts..."):
                cmp_open = fetch_cmp(tuple(odf["script"].unique()))

        show = pd.DataFrame()
        show["Client"]    = odf["client"].apply(lambda c: CLIENT_NAMES.get(c,c))
        show["Script"]    = odf["script"].values
        show["Entry"]     = odf["entry_date"].dt.strftime("%d/%m/%Y")
        show["Qty"]       = odf["buy_qty"].values
        show["Buy Price"] = odf["buy_price"].apply(lambda x: f"₹{x:,.2f}")
        show["Invested"]  = odf["invested"].apply(fmt)
        show["Days Held"] = odf["days_held"].values

        if cmp_open:
            show["CMP"]          = odf["script"].apply(lambda s: f"₹{cmp_open[s]:,.2f}" if cmp_open.get(s) else "—")
            show["Notional P&L"] = odf.apply(
                lambda r: fmt((cmp_open[r["script"]] - r["buy_price"]) * r["buy_qty"])
                          if cmp_open.get(r["script"]) else "—", axis=1).values
            show["Return %"]     = odf.apply(
                lambda r: f"{(cmp_open[r['script']]-r['buy_price'])/r['buy_price']*100:+.1f}%"
                          if cmp_open.get(r["script"]) else "—", axis=1).values
            st.dataframe(styled(show, ["Notional P&L"]), use_container_width=True, hide_index=True, height=520)
            total_notional = odf.apply(
                lambda r: (cmp_open[r["script"]] - r["buy_price"]) * r["buy_qty"]
                          if cmp_open.get(r["script"]) else 0, axis=1).sum()
            color = UP_COL if total_notional >= 0 else DN_COL
            st.markdown(f'<div style="text-align:right;font-size:.9rem;">Total Notional P&L: '
                        f'<b style="color:{color};">{fmt(total_notional)}</b></div>',
                        unsafe_allow_html=True)
        else:
            st.dataframe(show, use_container_width=True, hide_index=True, height=520)
            st.caption("Click 'Fetch Live CMP' to see notional P&L for all positions")

        _xls_open = odf  # captured for combined export below

    def show_closed(cdf, tab_key):
        # ── Excel-style column filters ──
        cf1, cf2, cf3 = st.columns([2,2,2])
        f_cli  = cf1.multiselect("Client", options=sorted(cdf["client"].unique()),
                                 format_func=lambda c: CLIENT_NAMES.get(c,c),
                                 default=list(cdf["client"].unique()), key=f"fcli_{tab_key}")
        f_scr  = cf2.text_input("Script contains", key=f"fscr_{tab_key}")
        f_pnl  = cf3.selectbox("P&L filter", ["All","Winners only","Losers only"], key=f"fpnl_{tab_key}")

        cdf = filter_df(cdf.copy())
        if f_cli: cdf = cdf[cdf["client"].isin(f_cli)]
        if f_scr: cdf = cdf[cdf["script"].str.contains(f_scr.upper(), na=False)]
        if f_pnl == "Winners only": cdf = cdf[cdf["net_pnl"] > 0]
        if f_pnl == "Losers only":  cdf = cdf[cdf["net_pnl"] < 0]
        cdf = cdf.sort_values("exit_date", ascending=False)

        show = pd.DataFrame()
        show["Client"]    = cdf["client"].apply(lambda c: CLIENT_NAMES.get(c,c))
        show["Script"]    = cdf["script"].values
        show["Entry"]     = cdf["entry_date"].dt.strftime("%d/%m/%Y")
        show["Exit"]      = cdf["exit_date"].dt.strftime("%d/%m/%Y")
        show["Qty"]       = cdf["buy_qty"].values
        show["Buy"]       = cdf["buy_price"].apply(lambda x: f"₹{x:,.2f}")
        show["Sell"]      = cdf["sell_price"].apply(lambda x: f"₹{x:,.2f}")
        show["Gross P&L"] = cdf["pnl"].apply(fmt)
        show["Net P&L"]   = cdf["net_pnl"].apply(fmt)
        show["Net%"]      = cdf["net_ret_pct"].apply(lambda x: f"{x:+.1f}%")
        show["Days"]      = cdf["days_held"].values
        st.dataframe(styled(show,["Gross P&L","Net P&L"]), use_container_width=True, hide_index=True, height=460)
        return cdf

    with tab2:
        cdf_out = show_closed(deliv_df, "deliv")
        _xls_closed = cdf_out

    with tab3:
        # Intraday — consolidate by client + script + date to avoid FIFO split rows
        idf_raw = filter_df(intra_df.copy())
        if idf_raw.empty:
            st.info("No intraday trades in this filter range.")
        else:
            def wavg_buy(g):
                return round((g["buy_price"] * g["sell_qty"]).sum() / g["sell_qty"].sum(), 2) if g["sell_qty"].sum() else 0
            def wavg_sell(g):
                return round((g["sell_price"] * g["sell_qty"]).sum() / g["sell_qty"].sum(), 2) if g["sell_qty"].sum() else 0

            idf_grp = idf_raw.groupby(["client","script","entry_date"]).apply(
                lambda g: pd.Series({
                    "total_qty": g["sell_qty"].sum(),
                    "avg_buy":   wavg_buy(g),
                    "avg_sell":  wavg_sell(g),
                    "gross_pnl": g["pnl"].sum(),
                    "net_pnl":   g["net_pnl"].sum(),
                    "charges":   g["total_charges"].sum(),
                })
            ).reset_index().sort_values("entry_date", ascending=False)

            idf_show = idf_grp.copy()
            idf_show["client"]    = idf_show["client"].apply(lambda c: CLIENT_NAMES.get(c,c))
            idf_show["entry_date"]= idf_show["entry_date"].dt.strftime("%d/%m/%Y")
            idf_show["avg_buy"]   = idf_show["avg_buy"].apply(lambda x: f"₹{x:,.2f}")
            idf_show["avg_sell"]  = idf_show["avg_sell"].apply(lambda x: f"₹{x:,.2f}")
            idf_show["gross_pnl"] = idf_show["gross_pnl"].apply(fmt)
            idf_show["net_pnl"]   = idf_show["net_pnl"].apply(fmt)
            idf_show["charges"]   = idf_show["charges"].apply(fmt)
            idf_show.columns = ["Client","Script","Date","Qty","Avg Buy","Avg Sell","Gross P&L","Net P&L","Charges"]
            st.dataframe(styled(idf_show,["Gross P&L","Net P&L"]), use_container_width=True, hide_index=True, height=480)

            with st.expander("Show raw FIFO rows (individual matched trades)"):
                show_closed(intra_df, "intra_raw")

        _xls_intra = idf_raw if not idf_raw.empty else None

    # ── Single combined export for all position tabs ──
    xls = to_excel({
        "Open Positions":  _xls_open,
        "Closed Delivery": _xls_closed,
        "Intraday":        _xls_intra,
    })
    st.markdown("---")
    st.download_button("📥 Download Positions Report (Excel)", xls, "positions_report.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Stock Analysis":
# ═══════════════════════════════════════════════════════════════════════════════
    # ── stock selector ──
    all_scripts = sorted(df["script"].unique())
    sc1, sc2 = st.columns([3,1])
    sel_script = sc1.selectbox("🔎 Pick a stock to analyse", all_scripts, label_visibility="visible")
    with sc2:
        st.markdown("<br>", unsafe_allow_html=True)
        fetch_live = st.button("📡 Fetch Live CMP")

    sdf      = df[df["script"] == sel_script].copy()
    s_closed = sdf[sdf["position"]=="Closed"]
    s_open   = sdf[sdf["position"]=="Open"]
    s_intra  = sdf[sdf["intraday"]==True]
    s_deliv  = s_closed[s_closed["intraday"]==False]

    # ── live CMP ──
    cmp_val = None
    if fetch_live:
        with st.spinner("Fetching price..."):
            res = fetch_cmp((sel_script,))
            cmp_val = res.get(sel_script)
        if cmp_val:
            st.success(f"Live CMP: ₹{cmp_val:,.2f}")
        else:
            st.warning("Could not fetch CMP — try manually checking the .NS ticker.")

    st.markdown("---")

    # ── POSITION SNAPSHOT CARD ─────────────────────────────────────────────────
    total_open_qty    = s_open["buy_qty"].sum()
    avg_open_buy      = (s_open["buy_price"]*s_open["buy_qty"]).sum()/total_open_qty if total_open_qty else 0
    open_invested     = s_open["invested"].sum()
    realised_gross    = s_closed["pnl"].sum()
    realised_net      = s_closed["net_pnl"].sum()
    total_charges     = sdf["total_charges"].sum()
    unrealised        = (cmp_val - avg_open_buy)*total_open_qty if cmp_val and avg_open_buy else None

    st.markdown(f"### {sel_script} — Position Snapshot")
    col_a, col_b, col_c, col_d, col_e, col_f, col_g = st.columns(7)
    col_a.metric("Open Qty",          f"{total_open_qty:.0f}")
    col_b.metric("Avg Buy (open)",     f"₹{avg_open_buy:,.2f}" if avg_open_buy else "—")
    col_c.metric("Capital Deployed",   fmt(open_invested))
    col_d.metric("Realised Net P&L",   fmt(realised_net))
    col_e.metric("Intraday Net P&L",   fmt(s_intra["net_pnl"].sum()))
    col_f.metric("Total Charges Paid", fmt(total_charges))
    col_g.metric("Unrealised P&L",     fmt(unrealised) if unrealised is not None else "—",
                 delta=f"CMP ₹{cmp_val:,.2f}" if cmp_val else None)

    st.markdown("---")

    # ── TABS: Open | Closed (Delivery) | Intraday | Charts ──────────────────────
    t1, t2, t3, t4 = st.tabs([
        f"🟢 Open Positions ({len(s_open)})",
        f"📦 Closed – Delivery ({len(s_deliv)})",
        f"⚡ Intraday ({len(s_intra)})",
        f"📈 Charts & History",
    ])

    def client_badge(c):
        color = CLIENT_COLORS.get(c, "#888")
        name  = CLIENT_NAMES.get(c, c)
        return f'<span style="background:{color};color:white;padding:2px 7px;border-radius:10px;font-size:.75rem;font-weight:700;">{name}</span>'

    # ── TAB 1: Open ──
    with t1:
        if s_open.empty:
            st.info(f"No open positions in {sel_script} for selected clients.")
        else:
            # per-client open summary
            for c in sel_clients:
                co = s_open[s_open["client"]==c]
                if co.empty: continue
                color = CLIENT_COLORS.get(c,"#888")
                name  = CLIENT_NAMES.get(c,c)
                total_qty = co["buy_qty"].sum()
                avg_buy   = (co["buy_price"]*co["buy_qty"]).sum()/total_qty
                invested  = co["invested"].sum()
                unreal    = (cmp_val-avg_buy)*total_qty if cmp_val else None
                unreal_pct= (cmp_val-avg_buy)/avg_buy*100 if cmp_val and avg_buy else None

                st.markdown(f"""
                <div style="border-left:4px solid {color};padding:10px 16px;background:#1a1a2e;border-radius:6px;margin:8px 0;">
                  <b style="color:{color};">{name}</b> &nbsp;|&nbsp;
                  Qty: <b>{total_qty:.0f}</b> &nbsp;|&nbsp;
                  Avg Buy: <b>₹{avg_buy:,.2f}</b> &nbsp;|&nbsp;
                  Invested: <b>{fmt(invested)}</b>
                  {"&nbsp;|&nbsp; Unrealised: <b style='color:" + (UP_COL if unreal>=0 else DN_COL) + ";'>" + fmt(unreal) + f" ({unreal_pct:+.1f}%)</b>" if unreal is not None else ""}
                </div>""", unsafe_allow_html=True)

                # individual lots
                show = co[["entry_date","buy_qty","buy_price","invested","days_held"]].copy()
                show["entry_date"] = show["entry_date"].dt.strftime("%d/%m/%Y")
                show["buy_price"]  = show["buy_price"].apply(lambda x: f"₹{x:,.2f}")
                show["invested"]   = show["invested"].apply(fmt)
                if cmp_val:
                    show["Unrealised"] = co.apply(lambda r: fmt((cmp_val-r["buy_price"])*r["buy_qty"]), axis=1)
                show.columns = (["Entry","Qty","Buy Price","Invested","Days Held"] +
                                (["Unrealised P&L"] if cmp_val else []))
                st.dataframe(show, use_container_width=True, hide_index=True, height=180)

    # ── TAB 2: Closed Delivery ──
    with t2:
        if s_deliv.empty:
            st.info(f"No closed delivery trades in {sel_script}.")
        else:
            # totals row
            tot_gross = s_deliv["pnl"].sum()
            tot_net   = s_deliv["net_pnl"].sum()
            tc1,tc2,tc3,tc4 = st.columns(4)
            tc1.metric("Trades",     len(s_deliv))
            tc2.metric("Gross P&L",  fmt(tot_gross))
            tc3.metric("Net P&L",    fmt(tot_net))
            tc4.metric("Charges",    fmt(s_deliv["total_charges"].sum()))

            show = s_deliv.sort_values("exit_date",ascending=False).copy()
            show["Client"]    = show["client"].apply(lambda c: CLIENT_NAMES.get(c,c))
            show["Entry"]     = show["entry_date"].dt.strftime("%d/%m/%Y")
            show["Exit"]      = show["exit_date"].dt.strftime("%d/%m/%Y")
            show["Qty"]       = show["buy_qty"].apply(lambda x: f"{x:.0f}")
            show["Buy ₹"]     = show["buy_price"].apply(lambda x: f"₹{x:,.2f}")
            show["Sell ₹"]    = show["sell_price"].apply(lambda x: f"₹{x:,.2f}")
            show["Gross P&L"] = show["pnl"].apply(fmt)
            show["Net P&L"]   = show["net_pnl"].apply(fmt)
            show["Return %"]  = show["net_ret_pct"].apply(lambda x: f"{x:+.1f}%")
            show["Days"]      = show["days_held"]
            show["Charges"]   = show["total_charges"].apply(fmt)
            cols_ = ["Client","Entry","Exit","Qty","Buy ₹","Sell ₹","Gross P&L","Net P&L","Return %","Days","Charges"]
            st.dataframe(styled(show[cols_],["Gross P&L","Net P&L"]),
                         use_container_width=True, hide_index=True, height=400)

    # ── TAB 3: Intraday ──
    with t3:
        if s_intra.empty:
            st.info(f"No intraday trades in {sel_script}.")
        else:
            tc1,tc2,tc3,tc4 = st.columns(4)
            tc1.metric("Intraday Days", s_intra["entry_date"].dt.date.nunique())
            tc2.metric("Total Qty Traded", f"{s_intra['sell_qty'].sum():.0f}")
            tc3.metric("Gross P&L", fmt(s_intra["pnl"].sum()))
            tc4.metric("Net P&L",   fmt(s_intra["net_pnl"].sum()))

            # consolidated by client + date
            def wavg(g, price_col, qty_col):
                return round((g[price_col]*g[qty_col]).sum()/g[qty_col].sum(),2) if g[qty_col].sum() else 0

            grp = s_intra.groupby(["client","entry_date"]).apply(
                lambda g: pd.Series({
                    "qty":       g["sell_qty"].sum(),
                    "avg_buy":   wavg(g,"buy_price","sell_qty"),
                    "avg_sell":  wavg(g,"sell_price","sell_qty"),
                    "gross_pnl": g["pnl"].sum(),
                    "net_pnl":   g["net_pnl"].sum(),
                    "charges":   g["total_charges"].sum(),
                })
            ).reset_index().sort_values("entry_date",ascending=False)

            grp["Client"]    = grp["client"].apply(lambda c: CLIENT_NAMES.get(c,c))
            grp["Date"]      = grp["entry_date"].dt.strftime("%d/%m/%Y")
            grp["Qty"]       = grp["qty"].apply(lambda x: f"{x:.0f}")
            grp["Avg Buy"]   = grp["avg_buy"].apply(lambda x: f"₹{x:,.2f}")
            grp["Avg Sell"]  = grp["avg_sell"].apply(lambda x: f"₹{x:,.2f}")
            grp["Gross P&L"] = grp["gross_pnl"].apply(fmt)
            grp["Net P&L"]   = grp["net_pnl"].apply(fmt)
            grp["Charges"]   = grp["charges"].apply(fmt)
            st.dataframe(styled(grp[["Client","Date","Qty","Avg Buy","Avg Sell","Gross P&L","Net P&L","Charges"]],
                                ["Gross P&L","Net P&L"]),
                         use_container_width=True, hide_index=True, height=400)

    # ── TAB 4: Charts ──
    with t4:
        if s_closed.empty:
            st.info("No closed trades to chart yet.")
        else:
            # Cumulative P&L line
            cum = s_closed.sort_values("exit_date").copy()
            cum["cum_net"]   = cum["net_pnl"].cumsum()
            cum["cum_gross"] = cum["pnl"].cumsum()
            cum["label"]     = (cum["exit_date"].dt.strftime("%d/%m/%y") + " · " +
                                cum["client"].apply(lambda c: CLIENT_NAMES.get(c,c)) + " " +
                                cum["buy_qty"].apply(lambda q: f"({q:.0f})"))
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=cum["label"], y=cum["cum_gross"],
                                     name="Cumulative Gross", mode="lines+markers",
                                     line=dict(color=ACC_COL,width=2)))
            fig.add_trace(go.Scatter(x=cum["label"], y=cum["cum_net"],
                                     name="Cumulative Net", mode="lines+markers",
                                     line=dict(color=UP_COL,width=2,dash="dot")))
            fig.add_hline(y=0, line_color=DN_COL, line_dash="dash", opacity=0.5)
            fig.update_layout(**CHART_THEME, height=300,
                              legend=dict(orientation="h",y=1.12))
            st.plotly_chart(fig, use_container_width=True)

            bar = s_closed.sort_values("exit_date").copy()
            bar["label"] = (bar["exit_date"].dt.strftime("%d/%m/%y") + " · " +
                            bar["client"].apply(lambda c: CLIENT_NAMES.get(c,c)))
            bar["color"] = bar["client"].apply(lambda c: CLIENT_COLORS.get(c,"#888"))
            fig2 = go.Figure(go.Bar(
                x=bar["label"], y=bar["net_pnl"], marker_color=bar["color"],
                marker_line_width=0,
                text=bar["net_pnl"].apply(lambda x: f"₹{x:,.0f}"), textposition="outside",
                hovertemplate="<b>%{x}</b><br>Net P&L: ₹%{y:,.0f}<extra></extra>"
            ))
            fig2.add_hline(y=0, line_color="#8890a8", line_dash="dash", opacity=0.4)
            fig2.update_layout(**CHART_THEME, height=280,
                               title="Net P&L per Trade · colored by client")
            st.plotly_chart(fig2, use_container_width=True)

    # ── Export ──
    xls_data = to_excel({
        "Open Positions": s_open  if not s_open.empty  else None,
        "Closed Trades":  s_deliv if not s_deliv.empty else None,
        "Intraday":       s_intra if not s_intra.empty else None,
    })
    st.download_button(f"📥 Download {sel_script} Report (Excel)", xls_data,
                       f"{sel_script}_report.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Brokerage":
# ═══════════════════════════════════════════════════════════════════════════════
    st.markdown("## Brokerage & Charges Analysis")

    # date range filter at top
    bc1, bc2 = st.columns(2)
    b_start = bc1.date_input("From", value=date(2025,1,1), key="bs")
    b_end   = bc2.date_input("To",   value=date.today(),   key="be")

    # filter by date on exit_date for closed, entry_date for open
    bdf = df.copy()
    ref_date = bdf.apply(lambda r: r["exit_date"] if pd.notna(r["exit_date"]) else r["entry_date"], axis=1)
    bdf = bdf[(ref_date.dt.date >= b_start) & (ref_date.dt.date <= b_end)]

    st.markdown("---")
    st.markdown('<div class="section-head">Charges Summary by Client</div>', unsafe_allow_html=True)

    rows = []
    for c in CLIENTS:
        if c not in sel_clients: continue
        cd = bdf[bdf["client"]==c]
        rows.append({
            "Client":        f"{CLIENT_NAMES.get(c,c)} ({c})",
            "# Trades":      len(cd),
            "Brokerage":     fmt(cd["buy_brokerage"].sum() + cd["sell_brokerage"].sum()),
            "STT":           fmt(cd["buy_stt"].sum()       + cd["sell_stt"].sum()),
            "GST":           fmt(cd["buy_gst"].sum()       + cd["sell_gst"].sum()),
            "Stamp Duty":    fmt(cd["buy_stamp"].sum()     + cd["sell_stamp"].sum()),
            "Txn Charges":   fmt(cd["buy_txn"].sum()       + cd["sell_txn"].sum()),
            "Total Charges": fmt(cd["total_charges"].sum()),
        })
    brows_df = pd.DataFrame(rows)
    st.dataframe(brows_df, use_container_width=True, hide_index=True)

    st.markdown("---")
    col1, col2 = st.columns(2)

    with col1:
        st.markdown('<div class="section-head">Charges Breakdown (all clients)</div>', unsafe_allow_html=True)
        tot = {
            "Brokerage":   bdf["buy_brokerage"].sum()+bdf["sell_brokerage"].sum(),
            "STT":         bdf["buy_stt"].sum()+bdf["sell_stt"].sum(),
            "GST":         bdf["buy_gst"].sum()+bdf["sell_gst"].sum(),
            "Stamp Duty":  bdf["buy_stamp"].sum()+bdf["sell_stamp"].sum(),
            "Txn Charges": bdf["buy_txn"].sum()+bdf["sell_txn"].sum(),
        }
        fig = px.pie(names=list(tot.keys()), values=list(tot.values()),
                     hole=0.5, color_discrete_sequence=[ACC_COL,"#a78bfa",UP_COL,"#fb923c",DN_COL],
                     height=280)
        fig.update_layout(**{k:v for k,v in CHART_THEME.items() if k not in ("xaxis","yaxis","margin")},
                          margin=dict(t=10,b=0))
        fig.update_traces(textinfo="label+percent", hovertemplate="%{label}: ₹%{value:,.0f}<extra></extra>")
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown('<div class="section-head">Monthly Charges vs Net P&L</div>', unsafe_allow_html=True)
        mdata = bdf.copy()
        mref = mdata.apply(lambda r: r["exit_date"] if pd.notna(r["exit_date"]) else r["entry_date"], axis=1)
        mdata["month"] = mref.dt.to_period("M").astype(str)
        mg = mdata.groupby("month").agg(charges=("total_charges","sum"), net_pnl=("net_pnl","sum")).reset_index()
        fig2 = go.Figure()
        fig2.add_trace(go.Bar(x=mg["month"], y=mg["net_pnl"], name="Net P&L",
                              marker_color=UP_COL, marker_line_width=0))
        fig2.add_trace(go.Bar(x=mg["month"], y=mg["charges"], name="Charges",
                              marker_color=DN_COL, marker_line_width=0))
        fig2.update_layout(**CHART_THEME, barmode="group", height=280)
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown("---")
    st.markdown('<div class="section-head">Intraday P&L Summary</div>', unsafe_allow_html=True)
    idf = bdf[bdf["intraday"]==True]
    if idf.empty:
        st.info("No intraday trades in this date range.")
    else:
        ic1,ic2,ic3,ic4 = st.columns(4)
        ic1.metric("Intraday Trades",  len(idf))
        ic2.metric("Gross P&L",        fmt(idf["pnl"].sum()))
        ic3.metric("Net P&L",          fmt(idf["net_pnl"].sum()))
        ic4.metric("Charges (intraday)", fmt(idf["total_charges"].sum()))

        ishow = idf[["client","script","entry_date","buy_qty","buy_price","sell_price","pnl","net_pnl"]].copy()
        ishow["entry_date"] = ishow["entry_date"].dt.strftime("%d/%m/%Y")
        ishow["buy_price"]  = ishow["buy_price"].apply(lambda x: f"₹{x:,.2f}")
        ishow["sell_price"] = ishow["sell_price"].apply(lambda x: f"₹{x:,.2f}")
        ishow["pnl"]        = ishow["pnl"].apply(fmt)
        ishow["net_pnl"]    = ishow["net_pnl"].apply(fmt)
        ishow.columns = ["Client","Script","Date","Qty","Buy","Sell","Gross P&L","Net P&L"]
        st.dataframe(styled(ishow,["Gross P&L","Net P&L"]), use_container_width=True, hide_index=True)

    xls = to_excel({
        "Brokerage by Client": brows_df,
        "Intraday Trades":     idf if not idf.empty else None,
    })
    st.download_button("📥 Download Brokerage Report (Excel)", xls, "brokerage_report.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Add / Close":
# ═══════════════════════════════════════════════════════════════════════════════
    tab1,tab2,tab3 = st.tabs(["🆕 Open Position","✅ Close Position","🗑️ Delete"])

    with tab1:
        with st.form("new"):
            c1,c2 = st.columns(2)
            client = c1.selectbox("Client", CLIENTS, format_func=lambda c: f"{CLIENT_NAMES.get(c,c)} ({c})")
            ptype  = c2.selectbox("Type", ["Long","Short","F&O"])
            script = st.text_input("Script").strip().upper()
            c3,c4,c5 = st.columns(3)
            edate = c3.date_input("Entry Date", value=date.today())
            qty   = c4.number_input("Buy Qty", min_value=1, step=1)
            price = c5.number_input("Buy Price ₹", min_value=0.01, step=0.05, format="%.2f")
            if st.form_submit_button("Add", type="primary"):
                if not script: st.error("Script required.")
                else:
                    trades = load()
                    nid = max((t.get("id",0) for t in trades), default=0)+1
                    t = {"id":nid,"client":client,"script":script,"type":ptype,
                         "entry_date":str(edate),"buy_qty":qty,"buy_price":price,
                         "buy_net_rate":price,
                         "exit_date":None,"sell_qty":0,"sell_price":0,"sell_net_rate":0,
                         "notes":"manual"}
                    for cc in CHARGE_COLS: t[cc]=0
                    trades.append(t); save(trades)
                    st.success(f"Added {script}"); st.balloons()

    with tab2:
        opts = {f"#{r['id']} | {CLIENT_NAMES.get(r['client'],r['client'])} | {r['script']} | {r['buy_qty']:.0f}@₹{r['buy_price']:,.2f}":r["id"]
                for _,r in open_df.iterrows()}
        if not opts: st.info("No open positions.")
        else:
            sel = st.selectbox("Select position", list(opts.keys()))
            sid = opts[sel]
            with st.form("close"):
                c1,c2,c3 = st.columns(3)
                xdate  = c1.date_input("Exit Date", value=date.today())
                sqty   = c2.number_input("Sell Qty", min_value=1, step=1)
                sprice = c3.number_input("Sell Price ₹", min_value=0.01, step=0.05, format="%.2f")
                if st.form_submit_button("Close", type="primary"):
                    trades = load()
                    for t in trades:
                        if t["id"]==sid:
                            pnl = (sqty*sprice)-(t["buy_qty"]*t["buy_price"])
                            pct = (sprice-t["buy_price"])/t["buy_price"]*100
                            t.update({"exit_date":str(xdate),"sell_qty":sqty,
                                      "sell_price":sprice,"sell_net_rate":sprice})
                            save(trades)
                            st.success(f"Closed — {fmt(pnl)} ({pct:+.1f}%)")
                            break

    with tab3:
        del_id = st.number_input("Trade ID to delete", min_value=1, step=1)
        if st.button("Delete", type="secondary"):
            trades = load()
            new = [t for t in trades if t.get("id")!=del_id]
            if len(new)<len(trades): save(new); st.success("Deleted"); st.rerun()
            else: st.error("Not found.")


# ═══════════════════════════════════════════════════════════════════════════════
elif page == "Settings":
# ═══════════════════════════════════════════════════════════════════════════════
    st.markdown("## ⚙️ Settings")

    # ── Import new broker CSVs ──────────────────────────────────────────────────
    st.markdown("### 📂 Import Broker CSVs")
    st.caption("Download the Trade Details CSV from Motilal Oswal → upload here → done. No terminal needed.")

    # dynamically import the processing logic from import_all
    import sys, importlib.util, io as _io

    @st.cache_resource
    def _load_import_mod():
        spec = importlib.util.spec_from_file_location(
            "import_all", os.path.join(os.path.dirname(DATA_FILE), "import_all.py"))
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return mod

    def _check_isin_conflicts(df):
        """Return info strings for ISINs that were auto-unified by ISIN remapping.
        These are handled automatically — no manual fix needed."""
        infos = []
        if 'ISIN' not in df.columns:
            return infos
        return infos  # conflicts are resolved before this check runs; nothing to report

    def _validate_fifo_balance(client, orders, client_trades):
        """Compare total CSV qty (all products) vs total FIFO output.
        Returns (errors, warnings). Errors block save; warnings are shown but allowed."""
        errors, warnings = [], []
        for scrip in sorted(orders['SCRIP'].unique()):
            sp = orders[orders['SCRIP'] == scrip]
            raw_buy  = round(float(sp[sp['SELL/BUY'] == 'B']['qty'].sum()), 2)
            raw_sell = round(float(sp[sp['SELL/BUY'] == 'S']['qty'].sum()), 2)
            ct = [t for t in client_trades if t['script'] == scrip]
            fifo_buy  = round(sum(t['buy_qty']  for t in ct), 2)
            fifo_sell = round(sum(t['sell_qty'] for t in ct), 2)
            diff_b = round(fifo_buy - raw_buy, 2)
            diff_s = round(fifo_sell - raw_sell, 2)
            # Sells with zero buy history = pre-existing position, just warn
            if raw_buy == 0 and raw_sell > 0 and fifo_sell == 0:
                warnings.append(
                    f"⚠️ `{client}` / **{scrip}**: {raw_sell:.0f} units sold — no buy history found (pre-existing position, skipped in FIFO)")
                continue
            # Over-match is always a bug
            if diff_b > 0.01:
                errors.append(f"❌ `{client}` / **{scrip}**: BUY qty mismatch — CSV={raw_buy:.0f} units, FIFO={fifo_buy:.0f} units (diff {diff_b:+.0f})")
            elif abs(diff_b) > 100.01:
                errors.append(f"❌ `{client}` / **{scrip}**: BUY qty mismatch — CSV={raw_buy:.0f} units, FIFO={fifo_buy:.0f} units (diff {diff_b:+.0f})")
            elif abs(diff_b) > 0.01:
                warnings.append(f"⚠️ `{client}` / **{scrip}**: BUY qty diff {diff_b:+.0f} (likely orphan intraday, ignored)")
            if diff_s > 0.01:
                # FIFO matched MORE sells than CSV — that's a real bug
                errors.append(f"❌ `{client}` / **{scrip}**: SELL qty mismatch — CSV={raw_sell:.0f} units, FIFO={fifo_sell:.0f} units (diff {diff_s:+.0f})")
            elif abs(diff_s) > 0.01:
                warnings.append(f"⚠️ `{client}` / **{scrip}**: SELL qty diff {diff_s:+.0f} (likely orphan intraday, ignored)")
        return errors, warnings

    def _parse_csv_to_orders(files, mod, already_processed_keys=None):
        """Load files, normalise, dedup by ORDER NO fingerprint, return sorted orders DataFrame.
        already_processed_keys: set of order keys already in the system — these rows are skipped.
        Returns (orders_df, new_keys_set, isin_conflict_warnings)
        """
        import pandas as pd
        already_processed_keys = already_processed_keys or set()
        frames = []
        for f in files:
            try: f.seek(0)
            except Exception: pass
            df_raw = pd.read_csv(f)
            df_raw['TRADE DATE'] = df_raw['TRADE DATE'].apply(mod.parse_date)
            df_raw['SELL/BUY']   = df_raw['SELL/BUY'].str.strip()
            df_raw['SCRIP NAME'] = df_raw['SCRIP NAME'].str.strip()
            df_raw['SCRIP']      = df_raw['SCRIP NAME'].apply(mod.norm)
            frames.append(df_raw)
        if not frames:
            return pd.DataFrame(), set(), []
        df = pd.concat(frames, ignore_index=True)
        # Cross-file dedup: same execution row in multiple uploaded files
        if 'TRADE NO' in df.columns:
            df = df.drop_duplicates(subset=['TRADE DATE', 'TRADE NO', 'SCRIP NAME', 'SELL/BUY', 'TRADE QTY'])

        # ISIN-based name unification — same ISIN = same stock regardless of BSE/NSE name variant.
        # When multiple SCRIP names map to the same ISIN, pick one canonical name and remap all rows.
        # This eliminates the need to manually add BSE/NSE name variants to import_all.py RAW dict.
        if 'ISIN' in df.columns:
            isin_remap = {}
            for isin, grp in df[df['ISIN'].notna() & (df['ISIN'].astype(str).str.strip() != '')].groupby('ISIN'):
                scrips = grp['SCRIP'].unique()
                if len(scrips) > 1:
                    # Prefer the name that norm() actually changed (exists in RAW dict = more canonical).
                    # Fall back to most frequent name in the data.
                    def _was_normalized(scrip, grp=grp):
                        sample_raw = grp[grp['SCRIP'] == scrip]['SCRIP NAME'].iloc[0]
                        return mod.norm(sample_raw.strip()) != sample_raw.strip()
                    normalized = [s for s in scrips if _was_normalized(s)]
                    if normalized:
                        canonical = normalized[0]
                    else:
                        counts = grp['SCRIP'].value_counts()
                        canonical = counts.index[0]
                    for s in scrips:
                        if s != canonical:
                            isin_remap[s] = canonical
            if isin_remap:
                df['SCRIP'] = df['SCRIP'].map(lambda s: isin_remap.get(s, s))

        # Skip rows from terminal 30023 (proprietary trades, not client trades)
        if 'TERMINAL NO' in df.columns:
            df = df[df['TERMINAL NO'].astype(str).str.strip() != '30023']

        # Dedup: remove exact duplicate rows (same file uploaded twice)
        # Only dedup by TRADE NO — never by price/qty alone (two orders can have same price+qty)
        before_dedup = len(df)
        if 'TRADE NO' in df.columns:
            df = df.drop_duplicates(subset=['TRADE DATE', 'TRADE NO', 'SCRIP NAME', 'SELL/BUY', 'TRADE QTY'])
        else:
            df = df.drop_duplicates()
        dupes_removed = before_dedup - len(df)
        if dupes_removed:
            st.info(f"ℹ️ Removed {dupes_removed} duplicate rows from uploaded file.")

        # Step 2: compute ORDER KEY fingerprint per raw row
        df['_order_key'] = df.apply(lambda r: make_order_key(
            r.get('ORDER NO',''), r.get('TRADE DATE',''),
            r.get('SCRIP NAME',''), r.get('SELL/BUY','')), axis=1)

        # Step 3: drop rows already processed in a previous import
        before       = len(df)
        new_keys     = set(df['_order_key'].tolist()) - already_processed_keys
        df           = df[df['_order_key'].isin(new_keys)]
        skipped      = before - len(df)

        charge_cols = ['BROKERAGE','TRANSACTION CHARGES','GST','STAMP DUTY',
                       'STT/CTT','SEBI CHARGES','IPFT CHARGES']
        for col in charge_cols:
            if col not in df.columns: df[col] = 0
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Normalise PRODUCT: treat any non-DELIVERY product as intraday
        if 'PRODUCT' not in df.columns:
            df['PRODUCT'] = 'DELIVERY'
        else:
            df['PRODUCT'] = df['PRODUCT'].str.strip().fillna('DELIVERY')
            # Only VALUE PLUS (MO intraday) is truly intraday — MTF, NORMAL, DELIVERY all carry over
            _INTRADAY = {'VALUE PLUS', 'MIS', 'INTRADAY', 'CO', 'BO'}
            df['PRODUCT'] = df['PRODUCT'].apply(lambda p: 'INTRADAY' if str(p).strip().upper() in _INTRADAY else 'DELIVERY')

        # ISIN conflict check — same ISIN resolving to different norm() names means a mapping gap
        isin_conflicts = _check_isin_conflicts(df)

        if df.empty:
            return pd.DataFrame(), new_keys, isin_conflicts

        # Fast vectorised aggregation — avoids slow groupby().apply() with lambda
        GRP = ['TRADE DATE','SCRIP','SELL/BUY','ORDER NO','PRODUCT']
        df['_amt'] = df['MARKET PRICE'] * df['TRADE QTY']
        agg = df.groupby(GRP, sort=False).agg(
            qty       =('TRADE QTY',          'sum'),
            _amt_sum  =('_amt',                'sum'),
            net_rate  =('NET RATE',            'mean'),
            brokerage =('BROKERAGE',           'sum'),
            stt       =('STT/CTT',             'sum'),
            gst       =('GST',                 'sum'),
            stamp     =('STAMP DUTY',          'sum'),
            txn_chrg  =('TRANSACTION CHARGES', 'sum'),
            _keys     =('_order_key',          lambda x: '||'.join(x)),
        ).reset_index()
        agg['price']    = (agg['_amt_sum'] / agg['qty']).round(2)
        agg['net_rate'] = agg['net_rate'].round(2)
        agg.drop(columns=['_amt_sum'], inplace=True)
        agg['_sort'] = pd.to_numeric(
            agg['ORDER NO'].astype(str).str.lstrip("'").str.strip(), errors='coerce').fillna(0)
        orders = agg.sort_values(['TRADE DATE','_sort']).drop(columns=['_sort'])
        return orders, new_keys, isin_conflicts

    def _fifo_from_queue(client, orders, buy_queue_init, trade_id_start):
        """Run FIFO matching. buy_queue_init: {(script,product): deque of buy dicts}
        DELIVERY and INTRADAY orders are kept in separate queues so they never mix."""
        from collections import deque
        CHARGE_K = ['brokerage','stt','gst','stamp','txn_chrg']
        all_trades, trade_id = [], trade_id_start
        # support old-style {script: deque} init from append mode (pre-PRODUCT)
        buy_queues = {}
        for k, q in buy_queue_init.items():
            key = k if isinstance(k, tuple) else (k, 'DELIVERY')
            buy_queues[key] = deque(list(q))

        prod_col = 'PRODUCT' if 'PRODUCT' in orders.columns else None

        for scrip in sorted(orders['SCRIP'].unique()):
            s = orders[orders['SCRIP']==scrip].sort_values(
                ['TRADE DATE', 'SELL/BUY'], ascending=[True, True]  # oldest first; B before S same day
            )
            for _, row in s.iterrows():
                prod = row[prod_col] if prod_col else 'DELIVERY'
                qkey = (scrip, prod)
                if qkey not in buy_queues:
                    buy_queues[qkey] = deque()
                bq = buy_queues[qkey]
                if row['SELL/BUY'] == 'B':
                    bq.append({k: row[k] for k in ['TRADE DATE','qty','price','net_rate']+CHARGE_K})
                else:
                    sell_rem = row['qty']
                    # Try primary queue first, then fall back to INTRADAY queue for
                    # DELIVERY sells (handles MO VALUE PLUS → DELIVERY conversion)
                    queues_to_try = [bq]
                    if prod == 'DELIVERY':
                        fb_q = buy_queues.get((scrip, 'INTRADAY'))
                        if fb_q: queues_to_try.append(fb_q)
                    for _bq in queues_to_try:
                        while sell_rem > 0 and _bq:
                            buy = _bq[0]
                            # Never match a buy dated after this sell — that's physically impossible
                            if buy['TRADE DATE'] > row['TRADE DATE']:
                                break
                            qty = min(buy['qty'], sell_rem)
                            fb  = qty/buy['qty'] if buy['qty'] else 0
                            fs  = qty/row['qty'] if row['qty'] else 0
                            t = {
                                'id': trade_id, 'client': client, 'script': scrip, 'type': 'Long',
                                'entry_date':    buy['TRADE DATE'].strftime('%Y-%m-%d'),
                                'buy_qty':       float(qty),
                                'buy_price':     float(buy['price']),
                                'buy_net_rate':  float(buy['net_rate']),
                                'exit_date':     row['TRADE DATE'].strftime('%Y-%m-%d'),
                                'sell_qty':      float(qty),
                                'sell_price':    float(row['price']),
                                'sell_net_rate': float(row['net_rate']),
                                'notes': 'imported',
                            }
                            for k in CHARGE_K:
                                t[f'buy_{k}']  = round(float(buy[k])*fb, 2)
                                t[f'sell_{k}'] = round(float(row[k])*fs, 2)
                            all_trades.append(t); trade_id += 1
                            buy['qty'] -= qty; sell_rem -= qty
                            if buy['qty'] <= 0: _bq.popleft()
                    # Unmatched sell qty = bought before CSV history — create break-even placeholder
                    if sell_rem > 0:
                        fs = sell_rem / row['qty'] if row['qty'] else 0
                        t = {
                            'id': trade_id, 'client': client, 'script': scrip, 'type': 'Long',
                            'entry_date':    row['TRADE DATE'].strftime('%Y-%m-%d'),
                            'buy_qty':       float(sell_rem),
                            'buy_price':     float(row['price']),
                            'buy_net_rate':  float(row['net_rate']),
                            'exit_date':     row['TRADE DATE'].strftime('%Y-%m-%d'),
                            'sell_qty':      float(sell_rem),
                            'sell_price':    float(row['price']),
                            'sell_net_rate': float(row['net_rate']),
                            'notes': 'pre-existing position — no buy history in CSV',
                        }
                        for k in CHARGE_K:
                            t[f'buy_{k}']  = 0
                            t[f'sell_{k}'] = round(float(row[k])*fs, 2)
                        all_trades.append(t); trade_id += 1

        # remaining open positions (only DELIVERY carries over; INTRADAY leftovers are MO's problem)
        for (scrip, prod), bq in buy_queues.items():
            if prod == 'INTRADAY':
                continue
            for buy in bq:
                t = {
                    'id': trade_id, 'client': client, 'script': scrip, 'type': 'Long',
                    'entry_date':    buy['TRADE DATE'].strftime('%Y-%m-%d'),
                    'buy_qty':       float(buy['qty']),
                    'buy_price':     float(buy['price']),
                    'buy_net_rate':  float(buy['net_rate']),
                    'exit_date': None, 'sell_qty': 0, 'sell_price': 0, 'sell_net_rate': 0,
                    'notes': 'imported',
                }
                for k in CHARGE_K:
                    t[f'buy_{k}']  = float(buy.get(k, 0))
                    t[f'sell_{k}'] = 0
                all_trades.append(t); trade_id += 1
        return all_trades

    def _run_import_from_uploads(uploaded_map: dict):
        """Full rebuild — process all CSVs from scratch. Resets processed order log."""
        import pandas as pd
        mod = _load_import_mod()
        all_trades, trade_id = [], 1
        new_processed = {}
        client_orders = {}  # pass to validation so it doesn't re-read CSVs
        for client, files in uploaded_map.items():
            orders, keys, _ = _parse_csv_to_orders(files, mod)  # no prior keys on full rebuild
            if orders.empty: continue
            client_orders[client] = orders
            new = _fifo_from_queue(client, orders, {}, trade_id)
            all_trades.extend(new)
            trade_id += len(new)
            new_processed[client] = keys
        save_processed_orders(new_processed)
        return all_trades, client_orders

    def _run_append_from_uploads(uploaded_map: dict):
        """Append mode — keep existing closed trades + historical open buy queues,
        process only genuinely NEW orders (by ORDER NO fingerprint). Safe to run daily."""
        import pandas as pd
        from collections import deque
        mod  = _load_import_mod()
        CHARGE_K = ['brokerage','stt','gst','stamp','txn_chrg']

        existing        = load()
        ex_df           = pd.DataFrame(existing) if existing else pd.DataFrame()
        result          = []
        max_id          = max((t.get('id',0) for t in existing), default=0) + 1
        processed_orders = load_processed_orders()  # {client: set(keys)}
        stats           = {}  # {client: {new, skipped}}
        client_new_keys = {}  # accumulated per-client new keys from first parse
        client_orders   = {}  # pass to validation so it doesn't re-read CSVs

        for client, files in uploaded_map.items():
            already_keys = processed_orders.get(client, set())
            orders, new_keys, _ = _parse_csv_to_orders(files, mod, already_keys)
            client_new_keys[client] = new_keys
            if not orders.empty:
                client_orders[client] = orders
            stats[client] = {'new': len(new_keys), 'skipped': len(already_keys)}

            if orders.empty:
                # nothing new — keep everything as-is for this client
                if not ex_df.empty and 'client' in ex_df.columns:
                    result.extend(ex_df[ex_df['client']==client].to_dict('records'))
                continue

            if not ex_df.empty and 'client' in ex_df.columns:
                c_trades = ex_df[ex_df['client']==client].copy()
                # keep ALL closed trades — dropping by date causes permanent data loss
                # because their order keys are already in processed_orders and can never be re-created
                keep_closed = c_trades[c_trades['exit_date'].notna()]
                result.extend(keep_closed.to_dict('records'))

                # rebuild buy queue from open positions that existed before new data
                open_positions = c_trades[c_trades['exit_date'].isna()]
                buy_queue_init = {}
                for _, r in open_positions.iterrows():
                    sc = r['script']
                    if sc not in buy_queue_init:
                        buy_queue_init[sc] = deque()
                    buy_queue_init[sc].append({
                        'TRADE DATE': pd.Timestamp(r['entry_date']),
                        'qty':        float(r['buy_qty']),
                        'price':      float(r['buy_price']),
                        'net_rate':   float(r.get('buy_net_rate', r['buy_price'])),
                        **{k: float(r.get(f'buy_{k}', 0)) for k in CHARGE_K}
                    })
            else:
                buy_queue_init = {}

            # run FIFO with existing open positions as starting queue
            new_trades = _fifo_from_queue(client, orders, buy_queue_init, max_id)

            # ── SAFETY NET: reject any impossible trade before it touches trades.json ──
            impossible = [
                t for t in new_trades
                if t.get('exit_date') and t['exit_date'] < t['entry_date']
            ]
            if impossible:
                msgs = [f"  • {t['script']} bought {t['entry_date']} sold {t['exit_date']}"
                        for t in impossible]
                raise ValueError(
                    f"FIFO produced {len(impossible)} impossible trade(s) for {client} "
                    f"(sold before bought):\n" + "\n".join(msgs) +
                    "\n\nThis usually means the CSV has an earlier sell with no matching buy. "
                    "Please upload the full trade history from the beginning."
                )

            result.extend(new_trades)
            max_id += len(new_trades)

        # keep other clients' trades untouched
        updated_clients = set(uploaded_map.keys())
        if not ex_df.empty and 'client' in ex_df.columns:
            other = ex_df[~ex_df['client'].isin(updated_clients)].to_dict('records')
            result.extend(other)

        # persist updated processed order keys — use keys from first parse, no second pass
        for client, new_keys in client_new_keys.items():
            processed_orders[client] = processed_orders.get(client, set()) | new_keys
        save_processed_orders(processed_orders)

        return result, stats, client_orders

    # ── CSV storage folder ──
    RAW_CSV_DIR = os.path.join(os.path.dirname(DATA_FILE), "raw_csvs")
    os.makedirs(RAW_CSV_DIR, exist_ok=True)

    def saved_csvs_for(client):
        d = os.path.join(RAW_CSV_DIR, client)
        if not os.path.exists(d): return []
        return sorted([os.path.join(d,f) for f in os.listdir(d) if f.endswith('.csv')])

    def save_uploaded_csv(client, uploaded_file):
        d = os.path.join(RAW_CSV_DIR, client)
        os.makedirs(d, exist_ok=True)
        dest = os.path.join(d, uploaded_file.name)
        already_exists = os.path.exists(dest)
        with open(dest, 'wb') as f:
            f.write(uploaded_file.getbuffer())
        return dest, already_exists

    # ── Upload UI ──
    uploaded_map = {}
    st.markdown("**Upload new CSVs** — already-saved files are reused automatically. Just upload from the date shown below:")

    # compute last trade date per client from trades.json
    _all_trades_df = pd.DataFrame(load())
    def _last_trade_date(client):
        if _all_trades_df.empty or 'client' not in _all_trades_df.columns:
            return None
        ct = _all_trades_df[_all_trades_df['client']==client]
        if ct.empty: return None
        raw = []
        for col in ['entry_date', 'exit_date']:
            if col in ct.columns:
                raw.extend(ct[col].dropna().tolist())
        dates = pd.to_datetime(raw, errors='coerce').dropna()
        return dates.max().strftime('%d %b %Y') if not dates.empty else None

    import_meta = load_import_meta()
    upcols = st.columns(len(CLIENTS))
    for i, c in enumerate(CLIENTS):
        with upcols[i]:
            color      = CLIENT_COLORS.get(c, "#818cf8")
            saved      = saved_csvs_for(c)
            last_trade = _last_trade_date(c)
            files_html = (f"<span style='color:#6b7280;font-size:.68rem;'>{len(saved)} file(s) saved</span>"
                          if saved else "<span style='color:#9ca3af;font-size:.68rem;'>No files saved yet</span>")
            if last_trade:
                date_html = (f"<div style='background:#1e3a2f;border:1px solid #34d399;border-radius:6px;"
                             f"padding:4px 8px;margin-bottom:6px;'>"
                             f"<span style='color:#6b7280;font-size:.68rem;'>Download from</span><br>"
                             f"<span style='color:#34d399;font-weight:700;font-size:.82rem;'>📅 {last_trade}</span>"
                             f"</div>")
            else:
                date_html = (f"<div style='background:#3a1e1e;border:1px solid #f87171;border-radius:6px;"
                             f"padding:4px 8px;margin-bottom:6px;'>"
                             f"<span style='color:#f87171;font-size:.72rem;'>No data yet — download full history</span>"
                             f"</div>")
            st.markdown(
                f"<div style='color:{color};font-weight:700;font-size:.82rem;margin-bottom:2px;'>"
                f"{CLIENT_NAMES[c]}</div>"
                f"<div style='color:#9ca3af;font-size:.72rem;margin-bottom:4px;'>{c}</div>"
                f"{date_html}"
                f"<div style='margin-bottom:6px;'>{files_html}</div>",
                unsafe_allow_html=True
            )
            up = st.file_uploader("Upload CSV", type="csv", accept_multiple_files=True,
                                  key=f"up_{c}", label_visibility="collapsed")
            if up:
                # ── Safety check: validate client code inside the CSV ──
                mismatch_files = []
                for uf in up:
                    try:
                        import io as _chk_io
                        peek = pd.read_csv(_chk_io.BytesIO(uf.read()), nrows=5)
                        uf.seek(0)  # reset after reading
                        # MO CSVs may have CLIENT CODE or ACCOUNT NO column
                        code_col = next((col for col in peek.columns
                                         if any(k in col.upper() for k in
                                                ['CLIENT','ACCOUNT','MEMBER'])), None)
                        if code_col:
                            codes_in_file = peek[code_col].dropna().astype(str).str.strip().unique()
                            if not any(c in code for code in codes_in_file for c in [c]):
                                # check if any OTHER client code appears
                                wrong = [code for code in codes_in_file
                                         if code.upper() != c.upper() and code != '']
                                if wrong:
                                    mismatch_files.append((uf.name, wrong[0]))
                    except Exception:
                        pass  # can't read → let import handle it

                if mismatch_files:
                    for fname, found_code in mismatch_files:
                        st.error(f"⛔ **Wrong file!** `{fname}` contains client `{found_code}` "
                                 f"but you're uploading to **{CLIENT_NAMES[c]} ({c})**. "
                                 f"Please upload the correct file.")
                else:
                    uploaded_map[c] = up

    st.markdown("<br>", unsafe_allow_html=True)

    # always show Run Import if any client has saved CSVs
    any_saved = any(saved_csvs_for(c) for c in CLIENTS)
    has_new   = bool(uploaded_map)

    if has_new:
        st.info(f"New files: {', '.join([f'{CLIENT_NAMES[c]} ({len(f)} file(s))' for c,f in uploaded_map.items()])} — will be saved and merged with existing history.")
    elif any_saved:
        st.info("No new files uploaded — will re-process all saved CSVs.")

    if has_new or any_saved:
        # mode selector
        import_mode = st.radio(
            "Import mode",
            ["➕ Append new data (recommended)", "🔄 Full rebuild from all saved CSVs"],
            captions=[
                "Keeps existing history, adds only new CSV data on top. Use this every month.",
                "Re-processes everything from scratch. Use only if history looks wrong."
            ],
            horizontal=True,
            key="import_mode"
        )

        btn_label = "🚀 Append" if "Append" in import_mode else "🚀 Full Rebuild"
        if st.button(btn_label, type="primary"):
            with st.spinner("Processing..."):
                try:
                    # save new uploads
                    overwritten = []
                    for c, files in uploaded_map.items():
                        for f in files:
                            _, existed = save_uploaded_csv(c, f)
                            if existed: overwritten.append(f.name)
                    if overwritten:
                        st.warning(f"⚠️ Overwritten (duplicates will be removed): {', '.join(overwritten)}")

                    import io as _io2

                    if "Append" in import_mode:
                        if not has_new:
                            st.warning("No new files uploaded — nothing to append.")
                            st.stop()
                        new_trades, append_stats, client_orders_map = _run_append_from_uploads(uploaded_map)
                        clients_updated = list(uploaded_map.keys())
                    else:
                        # Full Rebuild: wipe old saved files for each client that has
                        # new uploads, then save only the new ones.  This prevents
                        # the raw_csvs folder from accumulating overlapping files
                        # across sessions, which would cause double-counting in FIFO.
                        import shutil as _shutil
                        for c, files in uploaded_map.items():
                            old_dir = os.path.join(RAW_CSV_DIR, c)
                            if os.path.exists(old_dir):
                                _shutil.rmtree(old_dir)
                        # re-save only the freshly uploaded files
                        for c, files in uploaded_map.items():
                            for uf in files:
                                try: uf.seek(0)
                                except Exception: pass
                                save_uploaded_csv(c, uf)
                        full_map = {}
                        for c in CLIENTS:
                            paths = saved_csvs_for(c)
                            if paths:
                                full_map[c] = [_io2.BytesIO(open(p,'rb').read()) for p in paths]  # noqa: SIM115
                        new_trades, client_orders_map = _run_import_from_uploads(full_map)
                        clients_updated = list(full_map.keys())
                        append_stats = {}

                    # ── POST-FIFO INTEGRITY VALIDATION ──────────────────────────
                    # Only valid in full-rebuild mode: append mode has partial CSV data
                    # (today's orders only) while FIFO output includes historical positions,
                    # so qty comparisons would always produce false mismatches.
                    val_errors, val_warnings = [], []
                    if "Append" not in import_mode:
                        try:
                            for c_v in clients_updated:
                                full_orders_v = client_orders_map.get(c_v)
                                if full_orders_v is None or full_orders_v.empty:
                                    continue
                                c_trades_v = [t for t in new_trades if t.get('client') == c_v]
                                _v_errs, _v_warns = _validate_fifo_balance(c_v, full_orders_v, c_trades_v)
                                val_errors.extend(_v_errs)
                                val_warnings.extend(_v_warns)
                        except Exception as _ve:
                            val_warnings.append(f"⚠️ Validation could not complete: {_ve}")

                    if val_errors:
                        for _e in val_errors:
                            st.error(_e)
                        st.error("⛔ **FIFO balance errors — trades NOT saved.** The numbers above are wrong. "
                                 "Please share this screen with support before retrying.")
                        st.stop()

                    if val_warnings:
                        for _w in val_warnings:
                            st.warning(_w)
                    # ── END VALIDATION ───────────────────────────────────────────

                    save(new_trades)
                    closed  = [t for t in new_trades if t['exit_date']]
                    opened  = [t for t in new_trades if not t['exit_date']]
                    meta    = load_import_meta()
                    now_str = datetime.now(tz=__import__('zoneinfo').ZoneInfo('Asia/Kolkata')).strftime('%d %b %Y, %I:%M %p IST')
                    for c in clients_updated:
                        meta[c] = now_str
                    save_import_meta(meta)

                    if "Append" in import_mode:
                        total_new     = sum(s.get('new',0) for s in append_stats.values())
                        total_skipped = sum(s.get('skipped',0) for s in append_stats.values())
                        if total_new == 0:
                            st.info(f"✅ No new orders found — all {total_skipped} orders already in the system. Nothing changed.")
                        else:
                            st.success(f"✅ Appended {total_new} new orders · {total_skipped} duplicates skipped safely · {len(opened)} open, {len(closed)} closed")
                            st.balloons()
                    else:
                        st.success(f"✅ Full rebuild — {len(new_trades)} trades ({len(opened)} open, {len(closed)} closed)")
                        st.balloons()

                    st.cache_data.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"Import failed: {e}")

        # show saved files with delete option
        with st.expander("🗂️ Manage saved CSV files"):
            for c in CLIENTS:
                saved = saved_csvs_for(c)
                if saved:
                    st.markdown(f"**{CLIENT_NAMES[c]}** ({c})")
                    for p in saved:
                        col1, col2 = st.columns([5,1])
                        col1.caption(os.path.basename(p))
                        if col2.button("🗑️", key=f"del_{p}"):
                            os.remove(p)
                            st.rerun()
    else:
        st.caption("↑ Upload CSVs above to get started.")

    st.markdown("---")

    # ── Auto-schedule ──
    st.markdown("### 🗓️ Daily Auto-import (Windows Task Scheduler)")
    st.markdown("Set this up once — it watches your Downloads folder and re-imports whenever new CSVs appear:")

    bat_content = """@echo off
REM Auto-import: runs every morning at 8am via Task Scheduler
cd /d "%~dp0"
python import_all.py >> import_log.txt 2>&1
echo Import done at %date% %time% >> import_log.txt
"""

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**Step 1** — Open Task Scheduler (search in Start Menu)")
        st.markdown("**Step 2** — Create Basic Task → name it `Client Tracker Import`")
        st.markdown("**Step 3** — Trigger: Daily at 8:00 AM")
        st.markdown("**Step 4** — Action: Start a program → point to `auto_import.bat` in your project folder")
        st.markdown("**Step 5** — Finish. Done forever.")
    with col2:
        st.code(bat_content, language="bat")
        st.download_button("📥 Download auto_import.bat", bat_content,
                           "auto_import.bat", "text/plain")

    st.info("💡 **Workflow**: Each morning, download the CSV from MO → save to Downloads folder → the scheduled task re-imports automatically at 8am. Or just upload above manually anytime.")

    st.markdown("---")

    # ── Google Sheets Sync ──────────────────────────────────────────────────────
    st.markdown("### 📊 Sync to Google Sheets")
    st.markdown("Pushes all trade data to your shared Google Sheet — your employee just opens the link.")

    sheet_url = f"https://docs.google.com/spreadsheets/d/{GSHEET_ID}"
    st.markdown(f"🔗 **[Open Google Sheet]({sheet_url})**", unsafe_allow_html=False)

    col_sync1, col_sync2 = st.columns([1, 3])
    with col_sync1:
        if st.button("🔄 Sync Now", type="primary"):
            trades_data = load()
            with st.spinner("Pushing to Google Sheets..."):
                try:
                    result = sync_to_gsheet(trades_data)
                    st.success(result)
                except Exception as e:
                    err = str(e)
                    if "PERMISSION_DENIED" in err or "403" in err:
                        st.error("❌ Permission denied — make sure you shared the sheet with:\n`client-tracker-mofsl@client-tracker-mofsl.iam.gserviceaccount.com`")
                    else:
                        st.error(f"❌ Sync failed: {err}")
    with col_sync2:
        st.caption("Updates 5 tabs: Overview · Open Positions · Closed Trades · P&L Summary · Trade Ledger\nShare the sheet link with your employee — they get read-only view.")

    st.markdown("---")

    st.markdown("---")

    # ── Ticker Corrections ─────────────────────────────────────────────────────
    st.markdown("### 🔧 Fix Script Tickers (for CMP)")
    st.caption("Scripts below have no live price. Enter the correct NSE ticker (e.g. ALKYLAMINE, HDFCBANK, INFY). "
               "Find it at [finance.yahoo.com](https://finance.yahoo.com) — search the stock and copy the part before `.NS`.")

    trades_for_tickers = load()
    df_tk = pd.DataFrame(trades_for_tickers)
    overrides = load_ticker_overrides()

    if not df_tk.empty:
        open_scripts = df_tk[df_tk['exit_date'].isna()]['script'].unique().tolist()
        all_scripts  = sorted(df_tk['script'].unique().tolist())

        # find which ones have no CMP
        if open_scripts:
            test_cmp = fetch_cmp(tuple(open_scripts))
            missing  = sorted([s for s, v in test_cmp.items() if not v])
            working  = sorted([s for s, v in test_cmp.items() if v])
        else:
            missing, working = [], []

        if missing:
            st.markdown(f"**{len(missing)} script(s) with no CMP** — enter correct NSE ticker and click Save All:")
            st.caption("Tip: go to finance.yahoo.com → search stock → ticker shown is e.g. `ALKYLAMINE.NS` → enter just `ALKYLAMINE`")
            st.markdown("")
            cols_per_row = 3
            for i in range(0, len(missing), cols_per_row):
                row_scripts = missing[i:i+cols_per_row]
                cols = st.columns(cols_per_row)
                for j, sc in enumerate(row_scripts):
                    with cols[j]:
                        # show current override or SYMBOL_MAP guess as placeholder, not value
                        current_guess = overrides.get(sc, SYMBOL_MAP.get(sc, ""))
                        st.text_input(
                            sc,
                            value=current_guess,
                            placeholder="e.g. ALKYLAMINE",
                            key=f"ticker_{sc}"
                        )

            st.markdown("")
            if st.button("💾 Save All Ticker Fixes", type="primary"):
                for sc in missing:
                    val = st.session_state.get(f"ticker_{sc}", "").strip().upper().replace(".NS","")
                    if val:
                        overrides[sc] = val
                save_ticker_overrides(overrides)
                st.cache_data.clear()
                st.session_state["ticker_save_ok"] = True
                st.rerun()

            if st.session_state.pop("ticker_save_ok", False):
                st.success(f"✅ Saved {len(missing)} ticker fixes! CMP will now load for these scripts.")

        else:
            st.success(f"✅ All {len(open_scripts)} open position scripts have live CMP!")

        if working:
            with st.expander(f"✅ {len(working)} scripts already have live CMP"):
                st.write(", ".join(working))

        # edit any script
        st.markdown("")
        with st.expander("🔍 Edit ticker for any script"):
            sc2 = st.selectbox("Script", options=all_scripts, key="ticker_any_script")
            if sc2:
                cur2 = overrides.get(sc2, SYMBOL_MAP.get(sc2, ""))
                new2 = st.text_input("NSE Ticker", value=cur2,
                                      placeholder="e.g. HDFCBANK",
                                      key="ticker_any_val")
                if st.button("Save", key="save_any_ticker"):
                    overrides[sc2] = new2.strip().upper().replace(".NS","")
                    save_ticker_overrides(overrides)
                    st.cache_data.clear()
                    st.session_state["any_ticker_save_ok"] = sc2
                    st.rerun()
            if st.session_state.pop("any_ticker_save_ok", None):
                st.success(f"✅ Saved!")
    else:
        st.info("No trades loaded yet.")

    st.markdown("---")
    st.markdown("### ⚠️ Danger Zone")
    if st.button("🗑️ Clear ALL trades", type="secondary"):
        if st.session_state.get("confirm_clear"):
            save([]); st.session_state["confirm_clear"]=False
            st.success("Cleared."); st.rerun()
        else:
            st.session_state["confirm_clear"]=True
            st.warning("Are you sure? Click again to permanently delete all trade data.")

# ═══════════════════════════════════════════════════════════════════════════════
elif page == "💰 Capital":
# ═══════════════════════════════════════════════════════════════════════════════
    st.markdown("## 💰 Capital Tracker")
    st.caption("Track payins, payouts, and interest per client. Enter manually — takes 2 minutes per month.")

    ledger_data = load_ledger()

    # ── Add Entry Form ────────────────────────────────────────────────────────
    st.markdown("### Add Entry")
    with st.form("add_ledger_entry", clear_on_submit=True):
        f1, f2, f3, f4, f5, f6 = st.columns([2, 1.5, 1.5, 1.5, 1.5, 2])
        f_client  = f1.selectbox("Client", CLIENTS, format_func=lambda c: f"{CLIENT_NAMES[c]} ({c})")
        f_date    = f2.date_input("Date", value=pd.Timestamp.today())
        f_type    = f3.selectbox("Type", ["Payin", "Payout", "Interest"])
        f_ledger  = f4.selectbox("Account", ["Delivery", "MTF"])
        f_amount  = f5.number_input("Amount (₹)", min_value=0.0, step=1000.0, format="%.2f")
        f_note    = f6.text_input("Note (optional)")
        submitted = st.form_submit_button("➕ Add", type="primary", use_container_width=True)

    if submitted:
        if f_amount <= 0:
            st.error("Amount must be greater than 0.")
        else:
            # Payin = positive, Payout = negative, Interest = negative (cost)
            signed = f_amount if f_type == "Payin" else -f_amount
            new_entry = {
                "date":      f_date.strftime("%Y-%m-%d"),
                "type":      f_type.upper(),
                "amount":    round(signed, 2),
                "ledger":    f_ledger,
                "narration": f_note.strip(),
            }
            existing = ledger_data.get(f_client, [])
            existing.append(new_entry)
            existing.sort(key=lambda x: x["date"])
            ledger_data[f_client] = existing
            save_ledger(ledger_data)
            st.success(f"✅ Added {f_type} ₹{f_amount:,.0f} for {CLIENT_NAMES[f_client]} on {f_date.strftime('%d %b %Y')}")
            st.rerun()

    st.markdown("---")

    # ── Summary cards ──────────────────────────────────────────────────────────
    st.markdown("### Capital Summary")

    trades_for_cap = load()
    df_cap   = pd.DataFrame(trades_for_cap) if trades_for_cap else pd.DataFrame()
    open_cap = df_cap[df_cap["exit_date"].isna()] if not df_cap.empty else pd.DataFrame()
    cls_cap  = df_cap[df_cap["exit_date"].notna()] if not df_cap.empty else pd.DataFrame()

    total_payin = total_payout = total_interest = total_deployed = total_booked = 0
    cap_rows = []
    for c in CLIENTS:
        entries_c = ledger_data.get(c, [])
        payin    = sum(e["amount"] for e in entries_c if e["type"] == "PAYIN")
        payout   = sum(-e["amount"] for e in entries_c if e["type"] == "PAYOUT")
        interest = sum(-e["amount"] for e in entries_c if e["type"] == "INTEREST")
        net_cap  = payin - payout
        deployed = 0
        if not open_cap.empty:
            cc = open_cap[open_cap["client"] == c]
            deployed = (cc["buy_price"] * cc["buy_qty"]).sum() if not cc.empty else 0
        booked = cls_cap[cls_cap["client"] == c]["net_pnl"].sum() if not cls_cap.empty else 0
        true_roi = (booked - interest) / net_cap * 100 if net_cap > 0 else None
        total_payin    += payin
        total_payout   += payout
        total_interest += interest
        total_deployed += deployed
        total_booked   += booked
        cap_rows.append({
            "Client":          CLIENT_NAMES[c],
            "Payin":           payin,
            "Payout":          payout,
            "Net Capital":     net_cap,
            "Interest Paid":   interest,
            "Deployed (Open)": deployed,
            "Booked Net P&L":  booked,
            "True ROI":        true_roi,
        })

    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("Total Payin",      f"₹{total_payin/1e5:.2f}L")
    mc2.metric("Total Payout",     f"₹{total_payout/1e5:.2f}L")
    mc3.metric("Net Capital In",   f"₹{(total_payin-total_payout)/1e5:.2f}L")
    mc4.metric("Interest Paid",    f"₹{total_interest/1e5:.2f}L",
               delta=f"True ROI {(total_booked-total_interest)/(total_payin-total_payout)*100:+.1f}%" if (total_payin-total_payout) > 0 else None)

    st.markdown("")
    cap_df = pd.DataFrame(cap_rows)
    def _fmt_c(v):
        if v is None or (isinstance(v, float) and pd.isna(v)): return "—"
        return f"₹{v:,.0f}"
    def _fmt_r(v):
        if v is None or (isinstance(v, float) and pd.isna(v)): return "—"
        return f"{v:+.1f}%"
    cap_show = cap_df.copy()
    for col in ["Payin","Payout","Net Capital","Interest Paid","Deployed (Open)","Booked Net P&L"]:
        cap_show[col] = cap_df[col].apply(_fmt_c)
    cap_show["True ROI"] = cap_df["True ROI"].apply(_fmt_r)
    st.dataframe(cap_show, use_container_width=True, hide_index=True)

    st.markdown("---")

    # ── Transaction History ────────────────────────────────────────────────────
    st.markdown("### Transaction History")
    det_col1, det_col2 = st.columns([2, 1])
    det_client = det_col1.selectbox("Client", CLIENTS,
                                    format_func=lambda c: f"{CLIENT_NAMES[c]} ({c})",
                                    key="led_detail_client")
    entries = ledger_data.get(det_client, [])

    if entries:
        det_df = pd.DataFrame(entries).sort_values("date", ascending=False).reset_index(drop=True)

        # Inline delete
        del_idx = det_col2.number_input("Delete row #", min_value=1,
                                         max_value=len(det_df), step=1,
                                         key="led_del_idx") - 1
        if det_col2.button("🗑️ Delete this row", key="led_del_btn"):
            sorted_entries = sorted(entries, key=lambda x: x["date"], reverse=True)
            sorted_entries.pop(int(del_idx))
            ledger_data[det_client] = sorted_entries
            save_ledger(ledger_data)
            st.rerun()

        det_df["amount_fmt"] = det_df["amount"].apply(lambda x: f"₹{abs(x):,.0f}")
        det_show = det_df[["date","type","ledger","amount_fmt","narration"]].copy()
        det_show.columns = ["Date","Type","Account","Amount","Note"]
        st.dataframe(det_show, use_container_width=True, hide_index=True, height=380)

        # Monthly breakdown chart
        st.markdown("#### Monthly Breakdown")
        det_df["month"] = det_df["date"].str[:7]
        _mgrp = det_df.groupby(["month","type"])["amount"].sum().reset_index()
        _mgrp["bar"] = _mgrp.apply(
            lambda r: r["amount"] if r["type"] == "PAYIN"
            else (-r["amount"] if r["type"] in ("PAYOUT","INTEREST") else r["amount"]), axis=1)
        _color_map = {"PAYIN": UP_COL, "PAYOUT": DN_COL, "INTEREST": ACC_COL}
        fig_led = px.bar(_mgrp, x="month", y="bar", color="type",
                         color_discrete_map=_color_map,
                         barmode="group", height=240,
                         labels={"month":"Month","bar":"₹","type":"Type"})
        fig_led.update_layout(**CHART_THEME)
        st.plotly_chart(fig_led, use_container_width=True)

        # Totals for selected client
        _ep = sum(e["amount"] for e in entries if e["type"]=="PAYIN")
        _eo = sum(-e["amount"] for e in entries if e["type"]=="PAYOUT")
        _ei = sum(-e["amount"] for e in entries if e["type"]=="INTEREST")
        tc1, tc2, tc3 = st.columns(3)
        tc1.metric("Total Payin",    f"₹{_ep:,.0f}")
        tc2.metric("Total Payout",   f"₹{_eo:,.0f}")
        tc3.metric("Interest Paid",  f"₹{_ei:,.0f}")
    else:
        st.info(f"No entries yet for {CLIENT_NAMES[det_client]}. Add one above.")

    # ── Danger zone ────────────────────────────────────────────────────────────
    st.markdown("---")
    with st.expander("⚠️ Clear all entries for a client"):
        cl_client = st.selectbox("Client", CLIENTS,
                                 format_func=lambda c: CLIENT_NAMES[c], key="led_clear_client")
        if st.button("🗑️ Clear", type="secondary", key="led_clear_btn"):
            if st.session_state.get("confirm_led_clear"):
                ledger_data[cl_client] = []
                save_ledger(ledger_data)
                st.success("Cleared.")
                st.session_state["confirm_led_clear"] = False
                st.rerun()
            else:
                st.session_state["confirm_led_clear"] = True
                st.warning("Click again to confirm.")
